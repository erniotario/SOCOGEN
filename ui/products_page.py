from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QSpinBox, QComboBox, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QMessageBox, QFrame, QDateEdit
)
from datetime import date
import sys
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QFont
from sqlalchemy import func

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    load_workbook = None
    OPENPYXL_AVAILABLE = False

from database import SessionLocal
from models import Product, ProductStock, StockEntry, StockOutput, Store
from ui.utils import hline, field, make_btn, page_header, panel, configure_table, status_label, set_status


_C = {
    "bg":       "#0D1117",
    "surface":  "#161B22",
    "surface2": "#1C2128",
    "border":   "#30363D",
    "border2":  "#21262D",
    "accent":   "#58A6FF",
    "accent2":  "#1F6FEB",
    "green":    "#3FB950",
    "red":      "#F85149",
    "orange":   "#D29922",
    "text":     "#E6EDF3",
    "muted":    "#7D8590",
    "label":    "#8B949E",
}

_FORM_STYLE = f"""
    QWidget#form_card {{
        background: {_C['surface']};
        border: 1px solid {_C['border']};
        border-radius: 10px;
    }}
    QLabel#form_title {{
        color: {_C['accent']};
        font-size: 11px;
        font-weight: bold;
        letter-spacing: 1px;
        background: transparent;
    }}
    QLabel#field_label {{
        color: {_C['label']};
        font-size: 11px;
        font-weight: 600;
        background: transparent;
    }}
    QLineEdit, QSpinBox, QComboBox, QDateEdit {{
        background: {_C['bg']};
        border: 1px solid {_C['border']};
        border-radius: 6px;
        color: {_C['text']};
        font-size: 13px;
        padding: 7px 10px;
        selection-background-color: {_C['accent2']};
    }}
    QLineEdit:focus, QSpinBox:focus, QComboBox:focus, QDateEdit:focus {{
        border: 1px solid {_C['accent']};
    }}
    QLineEdit::placeholder {{ color: {_C['muted']}; }}
    QComboBox::drop-down {{ border: none; width: 24px; }}
    QComboBox::down-arrow {{ image: none; border: none; }}
    QSpinBox::up-button, QSpinBox::down-button {{ width: 18px; }}
"""

_TABLE_STYLE = f"""
    QTableWidget {{
        background: {_C['bg']};
        gridline-color: {_C['border2']};
        color: {_C['text']};
        border: 1px solid {_C['border']};
        border-radius: 8px;
        font-size: 13px;
        outline: none;
    }}
    QTableWidget::item {{
        padding: 4px 8px;
        border: none;
    }}
    QTableWidget::item:selected {{
        background: #1C3461;
        color: {_C['text']};
    }}
    QHeaderView::section {{
        background: {_C['surface2']};
        color: {_C['label']};
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 0.8px;
        padding: 8px 10px;
        border: none;
        border-bottom: 2px solid {_C['accent2']};
        border-right: 1px solid {_C['border']};
    }}
    QHeaderView::section:last {{ border-right: none; }}
    QScrollBar:vertical {{
        background: {_C['surface']};
        width: 8px;
        border-radius: 4px;
    }}
    QScrollBar::handle:vertical {{
        background: {_C['border']};
        border-radius: 4px;
        min-height: 30px;
    }}
"""

_SEARCH_STYLE = f"""
    #search_container {{
        background: {_C['bg']};
        border: 1px solid {_C['border']};
        border-radius: 7px;
    }}
    #search_container:focus-within {{
        border-color: {_C['accent']};
    }}
"""

_BTN_PRIMARY = f"""
    QPushButton {{
        background: {_C['accent2']};
        color: white;
        border: none;
        border-radius: 6px;
        font-size: 12px;
        font-weight: 600;
        padding: 7px 14px;
    }}
    QPushButton:hover {{ background: #388BFD; }}
    QPushButton:pressed {{ background: #1158C7; }}
    QPushButton:disabled {{ background: {_C['border']}; color: {_C['muted']}; }}
"""

_BTN_SECONDARY = f"""
    QPushButton {{
        background: {_C['surface']};
        color: {_C['text']};
        border: 1px solid {_C['border']};
        border-radius: 6px;
        font-size: 12px;
        font-weight: 600;
        padding: 7px 14px;
    }}
    QPushButton:hover {{ background: {_C['surface2']}; border-color: {_C['accent']}; }}
    QPushButton:pressed {{ background: {_C['bg']}; }}
    QPushButton:disabled {{ color: {_C['muted']}; }}
"""

_BTN_DANGER = f"""
    QPushButton {{
        background: transparent;
        color: {_C['red']};
        border: 1px solid {_C['red']};
        border-radius: 6px;
        font-size: 12px;
        font-weight: 600;
        padding: 7px 14px;
    }}
    QPushButton:hover {{ background: rgba(248,81,73,0.12); }}
    QPushButton:pressed {{ background: rgba(248,81,73,0.22); }}
"""


def _sep():
    f = QFrame()
    f.setFrameShape(QFrame.VLine)
    f.setStyleSheet(f"color: {_C['border']}; max-width: 1px;")
    return f


class ProductsPage(QWidget):
    """
    Page produits — un produit est unique par référence (globalement).
    Le stock initial est géré par magasin via la table ProductStock.
    Le tableau affiche une ligne par (produit × magasin).
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.session = SessionLocal()
        self._all_rows = []          # [(prod_id, ref, des, unit, stock, store_name, ps_id, store_id), ...]
        self._edit_mode = False
        self._editing_product_id = None   # Product.id en cours d'édition
        self._editing_ps_id      = None   # ProductStock.id en cours d'édition
        self._form_visible = False
        self._build_ui()

    # ─────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        self.setStyleSheet(f"QWidget {{ background: {_C['bg']}; }}")

        root.addWidget(page_header("Produits", "Gérer le catalogue de produits"))
        root.addWidget(hline())

        content = QWidget()
        cl = QVBoxLayout(content)
        cl.setContentsMargins(28, 16, 28, 16)
        cl.setSpacing(12)
        root.addWidget(content)

        # ── Barre d'actions ───────────────────────────────────────────────
        action_bar = QHBoxLayout()
        action_bar.setSpacing(10)

        tbl_title = QLabel("CATALOGUE PRODUITS")
        tbl_title.setStyleSheet(f"""
            color: {_C['label']}; font-size: 11px; font-weight: 700;
            letter-spacing: 1px; background: transparent;
        """)
        action_bar.addWidget(tbl_title, alignment=Qt.AlignVCenter)
        action_bar.addStretch()

        search_container = QWidget()
        search_container.setObjectName("search_container")
        search_container.setStyleSheet(_SEARCH_STYLE)
        sc_layout = QHBoxLayout(search_container)
        sc_layout.setContentsMargins(10, 5, 10, 5)
        sc_layout.setSpacing(6)
        search_icon = QLabel("🔍")
        search_icon.setStyleSheet("background: transparent; font-size: 13px;")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Rechercher par référence, désignation ou magasin…")
        self.search_input.setMinimumWidth(300)
        self.search_input.setStyleSheet(
            f"QLineEdit {{ border: none; background: transparent; color: {_C['text']}; font-size: 13px; }}"
            f"QLineEdit::placeholder {{ color: {_C['muted']}; }}"
        )
        self.search_input.textChanged.connect(self._on_search)
        sc_layout.addWidget(search_icon)
        sc_layout.addWidget(self.search_input)
        action_bar.addWidget(search_container)
        action_bar.addWidget(_sep())

        self.btn_toggle_form = QPushButton("＋  Nouveau produit")
        self.btn_toggle_form.setStyleSheet(_BTN_PRIMARY)
        self.btn_toggle_form.clicked.connect(self._toggle_form)

        self.btn_import_excel = QPushButton("⬆  Importer Excel")
        self.btn_import_excel.setStyleSheet(_BTN_SECONDARY)
        self.btn_import_excel.clicked.connect(self._import_products)

        self.btn_export_pdf = QPushButton("⬇  Exporter PDF")
        self.btn_export_pdf.setStyleSheet(_BTN_SECONDARY)
        self.btn_export_pdf.clicked.connect(self._export_products_pdf)

        self.btn_delete = QPushButton("🗑  Supprimer")
        self.btn_delete.setStyleSheet(_BTN_DANGER)
        self.btn_delete.clicked.connect(self._delete_product)

        action_bar.addWidget(self.btn_toggle_form)
        action_bar.addWidget(self.btn_import_excel)
        action_bar.addWidget(self.btn_export_pdf)
        action_bar.addWidget(self.btn_delete)
        cl.addLayout(action_bar)

        # ── Tableau ───────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID", "RÉFÉRENCE", "DÉSIGNATION", "UNITÉ", "STOCK INITIAL (TOTAL)", "MAGASINS", "STOCK_ID"
        ])
        self.table.setStyleSheet(_TABLE_STYLE)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setShowGrid(True)
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setColumnWidth(1, 140)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 120)
        self.table.setColumnWidth(5, 150)
        self.table.setColumnHidden(0, True)   # ID masqué (usage interne)
        self.table.setColumnHidden(6, True)   # STOCK_ID masqué (usage interne)
        cl.addWidget(self.table, stretch=1)

        # ── Statut ────────────────────────────────────────────────────────
        self.status = QLabel("")
        self.status.setStyleSheet(f"""
            QLabel {{ color: {_C['muted']}; font-size: 12px;
                      padding: 4px 2px; background: transparent; }}
        """)
        cl.addWidget(self.status)

        # ── Formulaire ────────────────────────────────────────────────────
        self.form_card = QWidget()
        self.form_card.setObjectName("form_card")
        self.form_card.setStyleSheet(_FORM_STYLE)
        self.form_card.setVisible(False)

        fc_layout = QVBoxLayout(self.form_card)
        fc_layout.setContentsMargins(20, 18, 20, 18)
        fc_layout.setSpacing(14)

        self._form_title_label = QLabel("NOUVEAU PRODUIT")
        self._form_title_label.setObjectName("form_title")
        fc_layout.addWidget(self._form_title_label)

        sep_line = QFrame()
        sep_line.setFrameShape(QFrame.HLine)
        sep_line.setStyleSheet(f"color: {_C['border2']}; margin-bottom: 4px;")
        fc_layout.addWidget(sep_line)

        row1 = QHBoxLayout(); row1.setSpacing(16)
        self.f_reference   = QLineEdit(); self.f_reference.setPlaceholderText("REF-001")
        self.f_designation = QLineEdit(); self.f_designation.setPlaceholderText("Ex : Ciment Portland")
        self.f_unit        = QLineEdit(); self.f_unit.setPlaceholderText("Ex : sac, kg, litre")
        self.f_unit.setText("unité")
        row1.addWidget(self._labeled("Référence *", self.f_reference), 2)
        row1.addWidget(self._labeled("Désignation *", self.f_designation), 3)
        row1.addWidget(self._labeled("Unité", self.f_unit), 1)
        fc_layout.addLayout(row1)

        row2 = QHBoxLayout(); row2.setSpacing(16)
        self.f_store   = QComboBox()
        self.f_initial = QSpinBox()
        self.f_initial.setMaximum(9999999)
        self.f_initial.setValue(0)
        self.f_date = QDateEdit(QDate.currentDate())
        self.f_date.setCalendarPopup(True)
        self.f_date.setDisplayFormat("dd/MM/yyyy")
        row2.addWidget(self._labeled("Magasin *", self.f_store), 2)
        row2.addWidget(self._labeled("Stock initial", self.f_initial), 1)
        row2.addWidget(self._labeled("Date", self.f_date), 1)
        fc_layout.addLayout(row2)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_add = QPushButton("＋  Ajouter le produit")
        self.btn_add.setStyleSheet(_BTN_PRIMARY)
        self.btn_add.clicked.connect(self._add_product)

        self.btn_edit = QPushButton("✎  Modifier le produit")
        self.btn_edit.setStyleSheet(_BTN_SECONDARY)
        self.btn_edit.clicked.connect(self._edit_product)

        self.btn_cancel_edit = QPushButton("✕  Annuler")
        self.btn_cancel_edit.setStyleSheet(_BTN_DANGER)
        self.btn_cancel_edit.clicked.connect(self._cancel_edit)
        self.btn_cancel_edit.setVisible(False)

        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_edit)
        btn_row.addWidget(self.btn_cancel_edit)
        fc_layout.addLayout(btn_row)

        cl.addWidget(self.form_card)

        self._load_stores()
        self.refresh()

    # ── Toggle formulaire ─────────────────────────────────────────────────
    def _toggle_form(self):
        self._form_visible = not self._form_visible
        self.form_card.setVisible(self._form_visible)
        if self._form_visible:
            self.btn_toggle_form.setText("✕  Fermer")
            self.btn_toggle_form.setStyleSheet(_BTN_DANGER)
            self.f_reference.setFocus()
        else:
            self._reset_form_state()
            self.btn_toggle_form.setText("＋  Nouveau produit")
            self.btn_toggle_form.setStyleSheet(_BTN_PRIMARY)

    def _reset_form_state(self):
        self.f_reference.clear()
        self.f_designation.clear()
        self.f_unit.setText("unité")
        self.f_initial.setValue(0)
        self.f_date.setDate(QDate.currentDate())
        self.f_store.setCurrentIndex(0 if self.f_store.count() else -1)
        self._edit_mode          = False
        self._editing_product_id = None
        self._editing_ps_id      = None
        self.btn_add.setText("＋  Ajouter le produit")
        self.btn_add.setStyleSheet(_BTN_PRIMARY)
        self.btn_edit.setEnabled(True)
        self.btn_cancel_edit.setVisible(False)
        self._form_title_label.setText("NOUVEAU PRODUIT")

    def _reset_form(self):
        self._reset_form_state()
        self._form_visible = False
        self.form_card.setVisible(False)
        self.btn_toggle_form.setText("＋  Nouveau produit")
        self.btn_toggle_form.setStyleSheet(_BTN_PRIMARY)

    # ── Helpers UI ────────────────────────────────────────────────────────
    def _labeled(self, label_text: str, widget: QWidget) -> QWidget:
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        lay = QVBoxLayout(container)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(5)
        lbl = QLabel(label_text)
        lbl.setObjectName("field_label")
        lbl.setStyleSheet(
            f"color: {_C['label']}; font-size: 11px; font-weight: 600; background: transparent;"
        )
        lay.addWidget(lbl)
        lay.addWidget(widget)
        return container

    # ── Chargement magasins ───────────────────────────────────────────────
    def _load_stores(self):
        try:
            self.session.expire_all()
            current = self.f_store.currentData() if hasattr(self, "f_store") else None
            self.f_store.blockSignals(True)
            self.f_store.clear()
            for s in self.session.query(Store).order_by(Store.name).all():
                self.f_store.addItem(s.name, s.id)
            if current:
                for i in range(self.f_store.count()):
                    if self.f_store.itemData(i) == current:
                        self.f_store.setCurrentIndex(i)
                        break
            self.f_store.blockSignals(False)
        except Exception as e:
            print(f"[Products] _load_stores: {e}")

    # ── Refresh ───────────────────────────────────────────────────────────
    def refresh(self):
        self._load_stores()
        try:
            self.session.expire_all()

            # Récupérer tous les produits avec leurs stocks par magasin
            rows = (
                self.session.query(Product, ProductStock, Store.name)
                .outerjoin(ProductStock, ProductStock.product_id == Product.id)
                .outerjoin(Store, Store.id == ProductStock.store_id)
                .order_by(Product.reference, Store.name)
                .all()
            )

            # Calcul du stock réel : initial (ProductStock) + entrées - sorties
            init_sums = dict(
                self.session.query(ProductStock.product_id, func.sum(ProductStock.initial_stock))
                .group_by(ProductStock.product_id).all()
            )
            entry_sums = dict(
                self.session.query(StockEntry.reference, func.sum(StockEntry.quantity))
                .group_by(StockEntry.reference).all()
            )
            output_sums = dict(
                self.session.query(StockOutput.reference, func.sum(StockOutput.quantity))
                .group_by(StockOutput.reference).all()
            )

            # Grouper par produit : une seule ligne par produit
            from collections import OrderedDict
            grouped = OrderedDict()  # product_id -> dict
            for p, ps, store_name in rows:
                if p.id not in grouped:
                    init = init_sums.get(p.id, 0) or 0
                    current = init + entry_sums.get(p.reference, 0) - output_sums.get(p.reference, 0)
                    grouped[p.id] = {
                        "prod_id":   str(p.id),
                        "ref":       p.reference,
                        "des":       p.designation,
                        "unit":      p.unit,
                        "stock":     current,
                        "stores":    [],
                        "ps_id":     "",
                    }
                if ps:
                    if store_name:
                        grouped[p.id]["stores"].append(store_name)
                    if not grouped[p.id]["ps_id"]:
                        grouped[p.id]["ps_id"] = str(ps.id)

            self._all_rows = [
                (
                    g["prod_id"],
                    g["ref"],
                    g["des"],
                    g["unit"],
                    str(g["stock"]),
                    " / ".join(g["stores"]) if g["stores"] else "—",
                    g["ps_id"],
                )
                for g in grouped.values()
            ]
            self._render_rows(self._all_rows)
        except Exception as e:
            print(f"[Products] refresh: {e}")

    def _render_rows(self, rows):
        self.table.setRowCount(len(rows))
        for row_idx, cols in enumerate(rows):
            # Highlight rows with zero or negative stock in red
            try:
                is_low = int(cols[4]) <= 0
            except (ValueError, IndexError):
                is_low = False
            row_bg = QColor("#3D0000") if is_low else (QColor("#0D1F2D") if row_idx % 2 == 0 else QColor(_C["bg"]))
            for col, text in enumerate(cols):
                item = QTableWidgetItem(str(text))
                item.setBackground(row_bg)
                if col == 0:
                    item.setForeground(QColor(_C["muted"]))
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
                elif col == 1:
                    item.setForeground(QColor(_C["accent"]))
                    item.setFont(QFont("Segoe UI", 9, QFont.Bold))
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                elif col == 4:
                    item.setForeground(QColor(_C["red"] if is_low else _C["green"]))
                    item.setFont(QFont("Segoe UI", 9, QFont.Bold))
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                else:
                    item.setForeground(QColor(_C["text"]))
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.table.setItem(row_idx, col, item)
            self.table.setRowHeight(row_idx, 36)

    # ── Recherche ─────────────────────────────────────────────────────────
    def _on_search(self, text: str):
        query = text.strip().lower()
        if not query:
            self._render_rows(self._all_rows)
            return
        filtered = [
            r for r in self._all_rows
            if query in r[1].lower() or query in r[2].lower() or query in r[5].lower()
        ]
        self._render_rows(filtered)
        self._set_status(f"{len(filtered)} résultat(s) pour « {text.strip()} »", info=True)

    # ── Ajouter ───────────────────────────────────────────────────────────
    def _add_product(self):
        if self._edit_mode:
            return self._update_product()

        ref      = self.f_reference.text().strip()
        des      = self.f_designation.text().strip()
        unit     = self.f_unit.text().strip() or "unité"
        init     = self.f_initial.value()
        store_id = self.f_store.currentData()

        if not ref:
            self._set_status("La référence est obligatoire.", error=True); return
        if not des:
            self._set_status("La désignation est obligatoire.", error=True); return
        if store_id is None:
            self._set_status("Sélectionnez un magasin.", error=True); return

        try:
            # Chercher si le produit (référence) existe déjà
            product = self.session.query(Product).filter_by(reference=ref).first()

            if product is None:
                # Nouveau produit
                product = Product(reference=ref, designation=des, unit=unit)
                self.session.add(product)
                self.session.flush()  # pour obtenir product.id

            # Vérifier si le stock pour ce magasin existe déjà
            ps_existing = self.session.query(ProductStock).filter_by(
                product_id=product.id, store_id=store_id
            ).first()

            if ps_existing:
                store_name = self.f_store.currentText()
                self._set_status(
                    f"Le produit « {ref} » a déjà un stock dans le magasin {store_name}.",
                    error=True
                )
                self.session.rollback()
                return

            # Créer le stock pour ce magasin
            ps = ProductStock(product_id=product.id, store_id=store_id, initial_stock=init)
            self.session.add(ps)
            self.session.commit()

            store_name = self.f_store.currentText()
            self._reset_form()
            self._set_status(f"Produit « {des} » ajouté au magasin {store_name}.")
            self.refresh()

        except Exception as e:
            self.session.rollback()
            self._set_status(f"Erreur : {e}", error=True)

    # ── Modifier ──────────────────────────────────────────────────────────
    def _edit_product(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            self._set_status("Sélectionnez une ligne à modifier.", error=True)
            return

        row = rows[0].row()
        product_id = int(self.table.item(row, 0).text())
        ps_id_text = self.table.item(row, 6).text()  # STOCK_ID (col masquée)
        ps_id      = int(ps_id_text) if ps_id_text else None

        product = self.session.get(Product, product_id)
        if not product:
            self._set_status("Produit introuvable.", error=True)
            return

        self._edit_mode          = True
        self._editing_product_id = product_id
        self._editing_ps_id      = ps_id

        self.f_reference.setText(product.reference)
        self.f_designation.setText(product.designation)
        self.f_unit.setText(product.unit or "unité")
        self.f_date.setDate(QDate.currentDate())

        if ps_id:
            ps = self.session.get(ProductStock, ps_id)
            self.f_initial.setValue(ps.initial_stock if ps else 0)
            # Sélectionner le bon magasin
            for i in range(self.f_store.count()):
                if ps and self.f_store.itemData(i) == ps.store_id:
                    self.f_store.setCurrentIndex(i)
                    break
        else:
            self.f_initial.setValue(0)

        self.btn_add.setText("💾  Enregistrer les modifications")
        self.btn_add.setStyleSheet(_BTN_PRIMARY.replace(_C["accent2"], "#1E8449"))
        self.btn_edit.setEnabled(False)
        self.btn_cancel_edit.setVisible(True)
        self._form_title_label.setText("MODIFIER LE PRODUIT")

        if not self._form_visible:
            self._form_visible = True
            self.form_card.setVisible(True)
            self.btn_toggle_form.setText("✕  Fermer")
            self.btn_toggle_form.setStyleSheet(_BTN_DANGER)

        self._set_status(f"Modification du produit « {product.reference} » en cours.", info=True)

    def _update_product(self):
        if not self._editing_product_id:
            self._set_status("Aucun produit sélectionné pour la modification.", error=True)
            return

        ref      = self.f_reference.text().strip()
        des      = self.f_designation.text().strip()
        unit     = self.f_unit.text().strip() or "unité"
        init     = self.f_initial.value()
        store_id = self.f_store.currentData()

        if not ref:
            self._set_status("La référence est obligatoire.", error=True); return
        if not des:
            self._set_status("La désignation est obligatoire.", error=True); return
        if store_id is None:
            self._set_status("Sélectionnez un magasin.", error=True); return

        try:
            # Vérifier unicité référence (hors produit courant)
            existing = self.session.query(Product).filter_by(reference=ref).first()
            if existing and existing.id != self._editing_product_id:
                self._set_status(f"La référence « {ref} » existe déjà.", error=True)
                return

            product = self.session.get(Product, self._editing_product_id)
            if not product:
                self._set_status("Produit introuvable.", error=True)
                return

            product.reference   = ref
            product.designation = des
            product.unit        = unit

            # Mettre à jour ou créer le ProductStock
            if self._editing_ps_id:
                ps = self.session.get(ProductStock, self._editing_ps_id)
                if ps:
                    # Vérifier doublon si le magasin change
                    if ps.store_id != store_id:
                        conflict = self.session.query(ProductStock).filter_by(
                            product_id=product.id, store_id=store_id
                        ).first()
                        if conflict:
                            store_name = self.f_store.currentText()
                            self._set_status(
                                f"Le produit « {ref} » a déjà un stock dans {store_name}.",
                                error=True
                            )
                            return
                    ps.store_id      = store_id
                    ps.initial_stock = init
            else:
                # Pas de stock existant — en créer un
                ps = ProductStock(product_id=product.id, store_id=store_id, initial_stock=init)
                self.session.add(ps)

            self.session.commit()
            self._set_status(f"Produit « {des} » mis à jour.")
            self._reset_form()
            self.refresh()

        except Exception as e:
            self.session.rollback()
            self._set_status(f"Erreur : {e}", error=True)

    def _cancel_edit(self):
        self._reset_form()
        self._set_status("Modification annulée.", info=True)

    # ── Supprimer ─────────────────────────────────────────────────────────
    def _delete_product(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            self._set_status("Sélectionnez un produit à supprimer.", error=True)
            return

        row        = rows[0].row()
        product_id = int(self.table.item(row, 0).text())
        product_ref = self.table.item(row, 1).text()
        ps_id_text  = self.table.item(row, 6).text()
        ps_id       = int(ps_id_text) if ps_id_text else None
        store_name  = self.table.item(row, 5).text()

        # Compter combien de magasins ce produit a
        nb_stocks = self.session.query(ProductStock).filter_by(product_id=product_id).count()

        if nb_stocks > 1 and ps_id:
            # Proposer de supprimer seulement ce magasin ou tout le produit
            reply = QMessageBox.question(
                self, "Que supprimer ?",
                f"Le produit « {product_ref} » est présent dans {nb_stocks} magasin(s).\n\n"
                f"• Oui  → supprimer uniquement le stock du magasin « {store_name} »\n"
                f"• Non  → supprimer le produit entier (tous les magasins)",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            if reply == QMessageBox.Cancel:
                return
            if reply == QMessageBox.Yes:
                # Supprimer seulement ce ProductStock
                try:
                    ps = self.session.get(ProductStock, ps_id)
                    if ps:
                        self.session.delete(ps)
                        self.session.commit()
                        self._set_status(
                            f"Stock de « {product_ref} » dans « {store_name} » supprimé."
                        )
                        self.refresh()
                except Exception as e:
                    self.session.rollback()
                    self._set_status(f"Erreur : {e}", error=True)
                return

        # Supprimer le produit entier
        has_entries = self.session.query(StockEntry).filter_by(reference=product_ref).first()
        has_outputs = self.session.query(StockOutput).filter_by(reference=product_ref).first()

        if has_entries or has_outputs:
            reply = QMessageBox.question(
                self, "Confirmation",
                f"Le produit « {product_ref} » possède des mouvements de stock.\n"
                "Supprimer quand même (les entrées/sorties seront conservées) ?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        try:
            product = self.session.get(Product, product_id)
            if product:
                self.session.delete(product)
                self.session.commit()
                self._set_status(f"Produit « {product_ref} » supprimé.")
                self.refresh()
        except Exception as e:
            self.session.rollback()
            self._set_status(f"Erreur : {e}", error=True)

    # ── Import Excel ──────────────────────────────────────────────────────
    def _import_products(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Importer des produits Excel", "",
            "Excel Files (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(
                self, "Import impossible",
                f"Le module openpyxl est requis.\nInterpréteur: {sys.executable}"
            )
            return

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet    = workbook.active
            rows     = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Le fichier Excel est vide.")

            headers    = [str(c).strip().lower() if c else "" for c in rows[0]]
            column_map = {}
            for idx, h in enumerate(headers):
                if h in ("reference", "référence", "ref"):
                    column_map["reference"] = idx
                elif h in ("designation", "désignation", "description"):
                    column_map["designation"] = idx
                elif h in ("unite", "unité", "unit"):
                    column_map["unit"] = idx
                elif h in ("stock initial", "initial_stock", "initial stock"):
                    column_map["initial_stock"] = idx
                elif h in ("magasin", "store", "store name", "nom magasin"):
                    column_map["store"] = idx

            if "reference" not in column_map or "designation" not in column_map:
                raise ValueError("Colonnes obligatoires manquantes : référence et désignation.")

            default_store = self.session.query(Store).order_by(Store.name).first()
            imported = 0; skipped = 0

            for row in rows[1:]:
                if not row or row[column_map["reference"]] is None:
                    continue
                ref = str(row[column_map["reference"]]).strip()
                if not ref:
                    continue
                des = ""
                if row[column_map["designation"]] is not None:
                    des = str(row[column_map["designation"]]).strip()
                if not des:
                    continue

                unit = "unité"
                if "unit" in column_map and column_map["unit"] < len(row):
                    v = row[column_map["unit"]]
                    if v is not None:
                        unit = str(v).strip() or "unité"

                initial_stock = 0
                if "initial_stock" in column_map and column_map["initial_stock"] < len(row):
                    v = row[column_map["initial_stock"]]
                    if v is not None:
                        try:
                            initial_stock = int(v)
                        except Exception:
                            initial_stock = 0

                # Résoudre le magasin
                store_id = None
                if "store" in column_map and column_map["store"] < len(row):
                    v = row[column_map["store"]]
                    if v is not None:
                        sname = str(v).strip()
                        st = self.session.query(Store).filter(
                            func.lower(Store.name) == sname.lower()
                        ).first()
                        if st:
                            store_id = st.id
                if store_id is None and default_store:
                    store_id = default_store.id

                # Créer ou récupérer le produit
                product = self.session.query(Product).filter_by(reference=ref).first()
                if product is None:
                    product = Product(reference=ref, designation=des, unit=unit)
                    self.session.add(product)
                    self.session.flush()

                # Vérifier doublon stock (produit × magasin)
                if self.session.query(ProductStock).filter_by(
                    product_id=product.id, store_id=store_id
                ).first():
                    skipped += 1
                    continue

                ps = ProductStock(product_id=product.id, store_id=store_id, initial_stock=initial_stock)
                self.session.add(ps)
                imported += 1

            self.session.commit()
            self.refresh()
            msg = f"{imported} ligne(s) importée(s)."
            if skipped:
                msg += f" {skipped} doublon(s) ignoré(s)."
            self._set_status(msg)

        except Exception as e:
            self.session.rollback()
            print(f"[Products] Import Excel: {e}")
            QMessageBox.critical(self, "Erreur import", str(e))

    # ── Export PDF ────────────────────────────────────────────────────────
    def _export_products_pdf(self):
        if not self._all_rows:
            QMessageBox.information(self, "Aucune donnée", "Aucun produit à exporter.")
            return
        default = f"produits_{date.today().strftime('%Y%m%d')}.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter les produits en PDF", default, "PDF Files (*.pdf)"
        )
        if not path:
            return
        try:
            self._build_products_pdf(path, self._all_rows)
            QMessageBox.information(self, "Export réussi", f"Fichier PDF enregistré :\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export PDF", str(e))

    def _build_products_pdf(self, path: str, rows: list):
        import os
        from datetime import datetime
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
        )
        from models import CompanySettings

        cs = self.session.get(CompanySettings, 1)
        company_name    = (cs.name      if cs else None) or "SOCOGEN"
        company_address = (cs.address   if cs else None) or ""
        company_city    = (cs.city      if cs else None) or ""
        company_phone   = (cs.phone     if cs else None) or ""
        company_email   = (cs.email     if cs else None) or ""
        company_website = (cs.website   if cs else None) or ""
        company_tax_id  = (cs.tax_id    if cs else None) or ""
        company_rccm    = (cs.rccm      if cs else None) or ""
        logo_path       = (cs.logo_path if cs else None) or ""

        MARGIN = 15 * mm
        doc = SimpleDocTemplate(
            path, pagesize=landscape(A4),
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN, bottomMargin=MARGIN,
            title="Catalogue des produits",
        )

        C_PRIMARY    = colors.HexColor("#1A5276")
        C_PRIMARY_LT = colors.HexColor("#2E86C1")
        C_ACCENT     = colors.HexColor("#2874A6")
        C_WHITE      = colors.white
        C_HDR_BG     = colors.HexColor("#1A5276")
        C_ROW_ODD    = colors.HexColor("#EBF5FB")
        C_ROW_EVEN   = colors.white
        C_GREY_TXT   = colors.HexColor("#555555")
        C_BORDER     = colors.HexColor("#AEB6BF")
        C_BORDER_LT  = colors.HexColor("#D5D8DC")
        C_BAND_BG    = colors.HexColor("#EAF2F8")
        C_BLACK      = colors.HexColor("#0D1117")
        C_GREEN      = colors.HexColor("#1E8449")

        def ps(name, **kw): return ParagraphStyle(name, **kw)

        s_co_name  = ps("cn", fontSize=20, fontName="Helvetica-Bold", textColor=C_PRIMARY,  leading=24, spaceAfter=2)
        s_co_info  = ps("ci", fontSize=8,  fontName="Helvetica",      textColor=C_GREY_TXT, leading=12)
        s_co_legal = ps("cl", fontSize=7.5,fontName="Helvetica-Bold", textColor=C_PRIMARY,  leading=11, spaceBefore=3)
        s_doc_ttl  = ps("dt", fontSize=22, fontName="Helvetica-Bold", textColor=C_PRIMARY,  alignment=TA_RIGHT, leading=26)
        s_doc_sub  = ps("ds", fontSize=9,  fontName="Helvetica",      textColor=C_GREY_TXT, alignment=TA_RIGHT, leading=12)
        s_section  = ps("sc", fontSize=11, fontName="Helvetica-Bold", textColor=C_PRIMARY,  spaceBefore=8, spaceAfter=5)
        s_kpi_lbl  = ps("kl", fontSize=7.5,fontName="Helvetica",      textColor=C_GREY_TXT, leading=10, spaceAfter=2)
        s_kpi_val  = ps("kv", fontSize=14, fontName="Helvetica-Bold", textColor=C_BLACK,    leading=16)
        s_kpi_blu  = ps("kb", fontSize=14, fontName="Helvetica-Bold", textColor=C_ACCENT,   leading=16)
        s_th       = ps("th", fontSize=8.5,fontName="Helvetica-Bold", textColor=C_WHITE,    alignment=TA_CENTER, leading=11)
        s_th_l     = ps("tl", fontSize=8.5,fontName="Helvetica-Bold", textColor=C_WHITE,    alignment=TA_LEFT,   leading=11)
        s_td       = ps("td", fontSize=8,  fontName="Helvetica",      textColor=C_BLACK,    alignment=TA_LEFT,   leading=11)
        s_td_c     = ps("tc", fontSize=8,  fontName="Helvetica",      textColor=C_BLACK,    alignment=TA_CENTER, leading=11)
        s_td_id    = ps("ti", fontSize=8,  fontName="Helvetica",      textColor=C_GREY_TXT, alignment=TA_CENTER, leading=11)
        s_td_ref   = ps("tr", fontSize=8,  fontName="Helvetica-Bold", textColor=C_ACCENT,   alignment=TA_LEFT,   leading=11)
        s_td_stk   = ps("ts", fontSize=8,  fontName="Helvetica-Bold", textColor=C_GREEN,    alignment=TA_RIGHT,  leading=11)
        s_footer   = ps("ft", fontSize=7.5,fontName="Helvetica",      textColor=C_GREY_TXT, alignment=TA_CENTER, leading=10)
        s_footer_b = ps("fb", fontSize=7.5,fontName="Helvetica-Bold", textColor=C_PRIMARY,  alignment=TA_CENTER, leading=10)

        story = []

        # En-tête
        left_col = []
        if logo_path and os.path.exists(logo_path):
            try:
                left_col.append(Image(logo_path, width=45*mm, height=16*mm, kind="proportional"))
                left_col.append(Spacer(1, 3*mm))
            except Exception:
                pass
        left_col.append(Paragraph(company_name, s_co_name))
        info_parts = [x for x in [company_address, company_city] if x]
        if info_parts: left_col.append(Paragraph("  |  ".join(info_parts), s_co_info))
        contact = [x for x in [
            f"Tél : {company_phone}" if company_phone else "",
            f"Email : {company_email}" if company_email else "",
            company_website
        ] if x]
        if contact: left_col.append(Paragraph("  |  ".join(contact), s_co_info))
        legal = [x for x in [
            f"N° Contribuable : {company_tax_id}" if company_tax_id else "",
            f"RCCM : {company_rccm}" if company_rccm else ""
        ] if x]
        if legal:
            left_col.append(Spacer(1, 1*mm))
            left_col.append(Paragraph("    ".join(legal), s_co_legal))

        right_col = [
            Paragraph("CATALOGUE DES PRODUITS", s_doc_ttl),
            Spacer(1, 2*mm),
            Paragraph(f"Date d'édition : {datetime.now().strftime('%d/%m/%Y')}", s_doc_sub),
            Paragraph(f"Heure : {datetime.now().strftime('%H:%M')}", s_doc_sub),
            Spacer(1, 2*mm),
            Paragraph(f"{len(rows)} ligne(s) enregistrée(s)", s_doc_sub),
        ]
        hdr_tbl = Table([[left_col, right_col]], colWidths=[120*mm, None])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), C_WHITE),
            ("TOPPADDING",    (0,0),(-1,-1), 10),
            ("BOTTOMPADDING", (0,0),(-1,-1), 10),
            ("LEFTPADDING",   (0,0),(0,-1),  0),
            ("RIGHTPADDING",  (-1,0),(-1,-1),0),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("LINEABOVE",     (0,0),(-1,0),  5, C_PRIMARY),
            ("LINEBELOW",     (0,0),(-1,-1), 1.5, C_PRIMARY),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 5*mm))

        # KPI
        total_initial = sum(int(r[4]) for r in rows if str(r[4]).isdigit())
        stores_set    = set(r[5] for r in rows if r[5] and r[5] != "—")
        produits_set  = set(r[1] for r in rows)

        def kpi_cell(label, value, val_style):
            return [Paragraph(label.upper(), s_kpi_lbl), Paragraph(str(value), val_style)]

        kpi_data = [[
            kpi_cell("Références distinctes", str(len(produits_set)), s_kpi_val),
            kpi_cell("Stock initial total",   str(total_initial),     s_kpi_blu),
            kpi_cell("Magasins concernés",    str(len(stores_set)),   s_kpi_val),
            kpi_cell("Date d'export",         datetime.now().strftime("%d/%m/%Y"), s_kpi_val),
        ]]
        kpi_tbl = Table(kpi_data, colWidths=[55*mm, 55*mm, 55*mm, 55*mm])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), C_BAND_BG),
            ("BOX",           (0,0),(-1,-1), 1,   C_BORDER),
            ("LINEAFTER",     (0,0),(-2,-1), 0.5, C_BORDER_LT),
            ("TOPPADDING",    (0,0),(-1,-1), 10),
            ("BOTTOMPADDING", (0,0),(-1,-1), 10),
            ("LEFTPADDING",   (0,0),(-1,-1), 14),
            ("RIGHTPADDING",  (0,0),(-1,-1), 14),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("LINEBELOW",     (0,0),(-1,-1), 3, C_PRIMARY_LT),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 6*mm))

        # Tableau produits
        story.append(Paragraph("LISTE DES PRODUITS", s_section))
        story.append(Spacer(1, 2*mm))

        col_labels = ["ID", "RÉFÉRENCE", "DÉSIGNATION", "UNITÉ", "STOCK INITIAL (TOTAL)", "MAGASINS"]
        col_widths = [16*mm, 36*mm, 94*mm, 26*mm, 32*mm, 42*mm]

        s_th_c = ps("thc", fontSize=8.5, fontName="Helvetica-Bold",
                    textColor=C_WHITE, alignment=TA_CENTER, leading=11)

        header_row = [
            Paragraph(col_labels[0], s_th),
            Paragraph(col_labels[1], s_th_l),
            Paragraph(col_labels[2], s_th_l),
            Paragraph(col_labels[3], s_th_c),
            Paragraph(col_labels[4], s_th),
            Paragraph(col_labels[5], s_th_l),
        ]
        tbl_data = [header_row]
        for r in rows:
            p_id, p_ref, p_des, p_unit, p_stock, p_store = r[0], r[1], r[2], r[3], r[4], r[5]
            tbl_data.append([
                Paragraph(str(p_id),    s_td_id),
                Paragraph(str(p_ref),   s_td_ref),
                Paragraph(str(p_des),   s_td),
                Paragraph(str(p_unit),  s_td_c),
                Paragraph(str(p_stock), s_td_stk),
                Paragraph(str(p_store), s_td),
            ])

        prod_tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
        tbl_style = [
            ("BACKGROUND",    (0,0),(-1,0),  C_HDR_BG),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0),  8.5),
            ("TOPPADDING",    (0,0),(-1,0),  8),
            ("BOTTOMPADDING", (0,0),(-1,0),  8),
            ("FONTSIZE",      (0,1),(-1,-1), 8),
            ("TOPPADDING",    (0,1),(-1,-1), 5),
            ("BOTTOMPADDING", (0,1),(-1,-1), 5),
            ("LEFTPADDING",   (0,0),(-1,-1), 6),
            ("RIGHTPADDING",  (0,0),(-1,-1), 6),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("BOX",           (0,0),(-1,-1), 1,   C_BORDER),
            ("INNERGRID",     (0,0),(-1,-1), 0.3, C_BORDER_LT),
            ("LINEBELOW",     (0,0),(-1,0),  1.5, C_PRIMARY_LT),
        ]
        for i in range(1, len(rows) + 1):
            tbl_style.append(("BACKGROUND", (0,i),(-1,i), C_ROW_ODD if i%2==1 else C_ROW_EVEN))
        prod_tbl.setStyle(TableStyle(tbl_style))
        story.append(prod_tbl)

        # Pied de page
        story.append(Spacer(1, 6*mm))
        story.append(HRFlowable(width="100%", thickness=1, color=C_PRIMARY, spaceAfter=3*mm))
        footer_data = [[
            Paragraph(company_name, s_footer_b),
            Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", s_footer),
            Paragraph(f"{len(rows)} ligne(s)", s_footer),
        ]]
        footer_tbl = Table(footer_data, colWidths=[None, None, 45*mm])
        footer_tbl.setStyle(TableStyle([
            ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
            ("LEFTPADDING", (0,0),(-1,-1), 0),
            ("RIGHTPADDING",(0,0),(-1,-1), 0),
        ]))
        story.append(footer_tbl)
        doc.build(story)

    # ── Statut ────────────────────────────────────────────────────────────
    def _set_status(self, msg: str, error: bool = False, info: bool = False):
        if error:
            color, icon = _C["red"],   "✕  "
        elif info:
            color, icon = _C["accent"], "ℹ  "
        else:
            color, icon = _C["green"],  "✓  "
        self.status.setText(icon + msg)
        self.status.setStyleSheet(f"""
            QLabel {{
                color: {color}; font-size: 12px; font-weight: 600;
                padding: 4px 2px; background: transparent;
            }}
        """)
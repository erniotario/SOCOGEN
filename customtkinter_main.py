import os
import sys
import unicodedata
from datetime import date, datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import hashlib

import customtkinter as ctk
from sqlalchemy import func

from database import Base, engine, SessionLocal
from models import Product, Store, User, StockEntry, StockOutput, CompanySettings

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    load_workbook = None
    OPENPYXL_AVAILABLE = False

try:
    original_md5 = hashlib.md5
    def patched_md5(*args, **kwargs):
        kwargs.pop('usedforsecurity', None)
        return original_md5(*args, **kwargs)
    hashlib.md5 = patched_md5

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

Base.metadata.create_all(bind=engine)


class ProductManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SOCOGEN — Stock Manager")
        self.geometry("1280x820")
        self.minsize(1024, 700)
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")
        self.configure(fg_color="#0D1117")

        self.session = SessionLocal()
        self._current_page = 1
        self.editing_product_id = None
        self.editing_entry_id = None
        self.editing_output_id = None
        self.current_user = None
        self.current_role = None
        self._ensure_default_admin_user()
        self._build_ui()
        self._show_page(0)
        self._show_login_screen()

    def _build_ui(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=210, corner_radius=0, fg_color="#0F172A")
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(10, weight=1)

        logo = ctk.CTkLabel(self.sidebar, text="SOCOGEN", font=ctk.CTkFont(size=22, weight="bold"), text_color="#F0F9FF")
        logo.grid(row=0, column=0, pady=(24, 4), padx=16, sticky="w")
        subtitle = ctk.CTkLabel(self.sidebar, text="Stock Manager", font=ctk.CTkFont(size=12), text_color="#BAE6FD")
        subtitle.grid(row=1, column=0, padx=16, sticky="w")

        self.btn_dashboard = ctk.CTkButton(self.sidebar, text="Tableau de bord", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(0))
        self.btn_products = ctk.CTkButton(self.sidebar, text="Produits", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(1))
        self.btn_entries = ctk.CTkButton(self.sidebar, text="Entrées", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(2))
        self.btn_outputs = ctk.CTkButton(self.sidebar, text="Sorties", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(3))
        self.btn_reports = ctk.CTkButton(self.sidebar, text="Rapports", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(4))
        self.btn_transactions = ctk.CTkButton(self.sidebar, text="Transactions", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(5))
        self.btn_dashboard.grid(row=2, column=0, pady=(30, 8), padx=16, sticky="ew")
        self.btn_products.grid(row=3, column=0, pady=8, padx=16, sticky="ew")
        self.btn_entries.grid(row=4, column=0, pady=8, padx=16, sticky="ew")
        self.btn_outputs.grid(row=5, column=0, pady=8, padx=16, sticky="ew")
        self.btn_reports.grid(row=6, column=0, pady=8, padx=16, sticky="ew")
        self.btn_transactions.grid(row=7, column=0, pady=8, padx=16, sticky="ew")

        self.btn_settings = ctk.CTkButton(self.sidebar, text="Paramètres", corner_radius=8, height=30, fg_color="#0284C7", hover_color="#0369A1", text_color="#F0F9FF", font=ctk.CTkFont(size=11), command=lambda: self._show_page(6))
        self.btn_settings.grid(row=8, column=0, pady=8, padx=16, sticky="ew")
        self.btn_settings.configure(state="disabled")

        self.user_label = ctk.CTkLabel(self.sidebar, text="Utilisateur: invité", font=ctk.CTkFont(size=10), text_color="#F0F9FF")
        self.user_label.grid(row=9, column=0, padx=16, sticky="sw")

        version = ctk.CTkLabel(self.sidebar, text="v2.2.0", font=ctk.CTkFont(size=10), text_color="#BAE6FD")
        version.grid(row=10, column=0, pady=(0, 8), padx=16, sticky="sw")

        self.btn_logout = ctk.CTkButton(self.sidebar, text="Déconnexion", corner_radius=8, height=30, fg_color="#EF4444", hover_color="#DC2626", font=ctk.CTkFont(size=11), command=self._logout)
        self.btn_logout.grid(row=11, column=0, pady=(0, 16), padx=16, sticky="ew")

        self.content = ctk.CTkFrame(self, corner_radius=0, fg_color="#0D1117")
        self.content.grid(row=0, column=1, sticky="nsew", padx=(16, 16), pady=16)
        self.content.grid_rowconfigure(2, weight=1)
        self.content.grid_columnconfigure(0, weight=1)

    def _show_page(self, page_idx):
        self._current_page = page_idx
        for widget in list(self.content.winfo_children()):
            widget.destroy()
        self.content.grid_rowconfigure(2, weight=1)
        
        if page_idx == 6 and self.current_role != "admin":
            self._build_unauthorized_page()
            return

        if page_idx == 0:
            self._build_dashboard()
        elif page_idx == 1:
            self._build_products()
        elif page_idx == 2:
            self._build_entries()
        elif page_idx == 3:
            self._build_outputs()
        elif page_idx == 4:
            self._build_reports()
        elif page_idx == 5:
            self._build_transactions()
        elif page_idx == 6:
            self._build_settings()

    def _ensure_default_admin_user(self):
        admin = self.session.query(User).filter(User.role == "admin").first()
        if not admin:
            salt, password_hash = self._hash_password("admin123")
            admin = User(username="admin", password_salt=salt, password_hash=password_hash, role="admin")
            self.session.add(admin)
            self.session.commit()

    def _hash_password(self, password, salt=None):
        if salt is None:
            salt = os.urandom(16).hex()
        hashed = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 100000)
        return salt, hashed.hex()

    def _verify_password(self, password, salt, stored_hash):
        _, hashed = self._hash_password(password, salt)
        return hashed == stored_hash

    def _show_login_screen(self):
        if hasattr(self, "login_overlay") and self.login_overlay.winfo_exists():
            self.login_overlay.destroy()

        self.login_overlay = ctk.CTkFrame(self, fg_color="#111827")
        self.login_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)

        card = ctk.CTkFrame(self.login_overlay, corner_radius=16, fg_color="#111827", width=420, height=340)
        card.place(relx=0.5, rely=0.4, anchor="center")

        title = ctk.CTkLabel(card, text="Connexion SOCOGEN", font=ctk.CTkFont(size=22, weight="bold"))
        title.grid(row=0, column=0, columnspan=2, pady=(24, 8), padx=20)

        self.login_username = ctk.CTkEntry(card, placeholder_text="Nom d'utilisateur")
        self.login_username.grid(row=1, column=0, columnspan=2, padx=20, pady=(8, 8), sticky="ew")

        self.login_password = ctk.CTkEntry(card, placeholder_text="Mot de passe", show="*")
        self.login_password.grid(row=2, column=0, columnspan=2, padx=20, pady=(0, 12), sticky="ew")

        self.login_error_label = ctk.CTkLabel(card, text="", text_color="#DC2626")
        self.login_error_label.grid(row=3, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="ew")

        login_button = ctk.CTkButton(card, text="Se connecter", corner_radius=10, height=32, font=ctk.CTkFont(size=12, weight="bold"), command=self._attempt_login)
        login_button.grid(row=4, column=0, columnspan=2, padx=20, pady=(0, 12), sticky="ew")

        # Hint removed: do not expose admin credentials on the login screen

        self.login_username.bind("<Return>", lambda e: self.login_password.focus())
        self.login_password.bind("<Return>", lambda e: self._attempt_login())
        self.login_username.focus()

    def _attempt_login(self):
        username = self.login_username.get().strip()
        password = self.login_password.get().strip()
        if not username or not password:
            self.login_error_label.configure(text="Nom d'utilisateur et mot de passe requis.")
            return

        user = self.session.query(User).filter(func.lower(User.username) == username.lower()).first()
        if not user or not self._verify_password(password, user.password_salt, user.password_hash):
            self.login_error_label.configure(text="Identifiants invalides.")
            return

        self._on_login_success(user)

    def _on_login_success(self, user):
        self.current_user = user
        self.current_role = user.role
        self.user_label.configure(text=f"Connecté: {user.username} ({user.role})")
        self.btn_settings.configure(state="normal" if self.current_role == "admin" else "disabled")
        if hasattr(self, "login_overlay"):
            self.login_overlay.place_forget()
        self._show_page(0)

    def _logout(self):
        self.current_user = None
        self.current_role = None
        self.user_label.configure(text="Utilisateur: invité")
        self.btn_settings.configure(state="disabled")
        self._show_login_screen()

    def _build_unauthorized_page(self):
        title = ctk.CTkLabel(self.content, text="Accès refusé", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(24, 0), padx=16)
        message = ctk.CTkLabel(self.content, text="Vous devez être administrateur pour accéder à cette page.", font=ctk.CTkFont(size=12))
        message.grid(row=1, column=0, sticky="w", padx=16, pady=(12, 0))

    def _build_dashboard(self):
        title = ctk.CTkLabel(self.content, text="Tableau de bord", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Vue d'ensemble du stock", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        stats = [
            ("Produits", self.session.query(Product).count()),
            ("Magasins", self.session.query(Store).count()),
            ("Entrées", self.session.query(StockEntry).count()),
            ("Sorties", self.session.query(StockOutput).count()),
        ]
        stats_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        stats_frame.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        for idx, (label, value) in enumerate(stats):
            self._build_stat_card(stats_frame, label, value, 0, idx)

        latest_products = self.session.query(Product).order_by(Product.id.desc()).limit(5).all()
        latest_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        latest_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))
        latest_frame.grid_columnconfigure(0, weight=1)
        latest_title = ctk.CTkLabel(latest_frame, text="Derniers produits ajoutés", font=ctk.CTkFont(size=14, weight="bold"))
        latest_title.grid(row=0, column=0, sticky="w", padx=16, pady=(12, 0))

        if latest_products:
            for idx, product in enumerate(latest_products, start=1):
                row_text = f"{idx}. {product.reference} — {product.designation} ({product.unit})"
                item = ctk.CTkLabel(latest_frame, text=row_text, anchor="w")
                item.grid(row=idx, column=0, sticky="w", padx=16, pady=4)
        else:
            empty = ctk.CTkLabel(latest_frame, text="Aucun produit enregistré.", anchor="w")
            empty.grid(row=1, column=0, sticky="w", padx=16, pady=12)

    def _build_products(self):
        title = ctk.CTkLabel(self.content, text="Produits", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Gérer le catalogue de produits", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        form_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        form_panel.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        form_panel.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self.ref_entry = ctk.CTkEntry(form_panel, placeholder_text="Référence *")
        self.des_entry = ctk.CTkEntry(form_panel, placeholder_text="Désignation *")
        self.unit_entry = ctk.CTkEntry(form_panel, placeholder_text="Unité")
        self.unit_entry.insert(0, "unité")
        self.ref_entry.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.des_entry.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        self.unit_entry.grid(row=0, column=2, padx=10, pady=12, sticky="ew")

        self.store_combo = ctk.CTkComboBox(form_panel, values=self._get_store_names())
        self.store_combo.set(self._get_store_names()[0] if self._get_store_names() else "")
        self.stock_entry = ctk.CTkEntry(form_panel, placeholder_text="Stock initial")
        self.date_entry = ctk.CTkEntry(form_panel, placeholder_text="Date")
        self.date_entry.insert(0, date.today().strftime("%d/%m/%Y"))
        self.store_combo.grid(row=1, column=0, padx=10, pady=(0, 16), sticky="ew")
        self.stock_entry.grid(row=1, column=1, padx=10, pady=(0, 16), sticky="ew")
        self.date_entry.grid(row=1, column=2, padx=10, pady=(0, 16), sticky="ew")

        self.add_button = ctk.CTkButton(form_panel, text="+ Ajouter", corner_radius=10, height=28, font=ctk.CTkFont(size=11, weight="bold"), command=self.add_product)
        self.add_button.grid(row=0, column=3, padx=8, pady=10, sticky="ew", rowspan=2)

        self.cancel_button = ctk.CTkButton(form_panel, text="✕ Annuler", corner_radius=10, height=28, font=ctk.CTkFont(size=11, weight="bold"), command=self.cancel_edit_product, fg_color="#9CA3AF", hover_color="#6B7280")
        self.cancel_button.grid(row=0, column=4, padx=8, pady=10, sticky="ew", rowspan=2)

        self.save_button = ctk.CTkButton(form_panel, text="Enregistrer", corner_radius=10, height=28, font=ctk.CTkFont(size=11, weight="bold"), command=self.save_product, fg_color="#10B981", hover_color="#059669")
        self.save_button.grid(row=0, column=5, padx=8, pady=10, sticky="ew", rowspan=2)
        self.save_button.configure(state="disabled")

        control_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        control_panel.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 12))
        control_panel.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self.search_entry = ctk.CTkEntry(control_panel, placeholder_text="Rechercher produits...")
        self.search_entry.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.search_entry.bind("<KeyRelease>", lambda event: self.search_products())

        self.import_button = ctk.CTkButton(control_panel, text="Importer Excel", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.import_products)
        self.export_button = ctk.CTkButton(control_panel, text="Exporter PDF", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.export_products_pdf)
        self.delete_button = ctk.CTkButton(control_panel, text="Supprimer", corner_radius=8, height=28, font=ctk.CTkFont(size=11), fg_color="#E53E3E", hover_color="#D53F3F", command=self.delete_selected)
        self.modify_button = ctk.CTkButton(control_panel, text="Modifier", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.modify_selected_product)
        self.select_all_button = ctk.CTkButton(control_panel, text="Tout sélectionner", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.select_all_products)
        self.import_button.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        self.export_button.grid(row=0, column=2, padx=10, pady=12, sticky="ew")
        self.delete_button.grid(row=0, column=3, padx=10, pady=12, sticky="ew")
        self.modify_button.grid(row=0, column=4, padx=10, pady=12, sticky="ew")
        self.select_all_button.grid(row=0, column=5, padx=10, pady=12, sticky="ew")

        self.table_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        self.table_frame.grid(row=4, column=0, sticky="nsew", padx=16, pady=(0, 16))
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)

        self.table = ttk.Treeview(self.table_frame, columns=("id", "ref", "designation", "unit", "stock", "store"), show="headings", selectmode="extended")
        self.table.heading("id", text="ID")
        self.table.heading("ref", text="Référence")
        self.table.heading("designation", text="Désignation")
        self.table.heading("unit", text="Unité")
        self.table.heading("stock", text="Stock")
        self.table.heading("store", text="Magasin")
        self.table.column("id", width=40, anchor="center")
        self.table.column("ref", width=140, anchor="w")
        self.table.column("designation", width=320, anchor="w")
        self.table.column("unit", width=90, anchor="center")
        self.table.column("stock", width=90, anchor="center")
        self.table.column("store", width=160, anchor="w")
        self.table.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        self.table.bind("<Double-1>", self._edit_product_from_table)

        scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.table.yview)
        self.table.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=8)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview", background="#111827", foreground="#F8FAFC", fieldbackground="#111827", rowheight=30, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#0F172A", foreground="#F8FAFC", relief="flat")
        style.map("Treeview", background=[("selected", "#4F8CFF")], foreground=[("selected", "white")])
        self.table.tag_configure("oddrow", background="#111827")
        self.table.tag_configure("evenrow", background="#0F172A")
        self.table.tag_configure("lowstock", background="#7F1D1D", foreground="#FCA5A5")

        self.status_label = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.status_label.grid(row=5, column=0, padx=16, sticky="w")

        self.refresh_products()

    def _build_entries(self):
        title = ctk.CTkLabel(self.content, text="Entrées", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Gérer les entrées de stock", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        form_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        form_panel.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        form_panel.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self.entry_product_values = self._get_product_references()
        self.entry_product_var = tk.StringVar(value=self.entry_product_values[0] if self.entry_product_values else "")
        self.entry_product_combo = ttk.Combobox(form_panel, textvariable=self.entry_product_var, values=self.entry_product_values, state="normal")
        if self.entry_product_values:
            self.entry_product_combo.set(self.entry_product_values[0])
        self.entry_product_combo.bind("<<ComboboxSelected>>", lambda e: self.on_entry_product_selected(self.entry_product_var.get()))
        self.entry_product_combo.bind("<KeyRelease>", self._on_entry_product_typed)
        self.entry_date = ctk.CTkEntry(form_panel, placeholder_text="Date (jj/mm/aaaa)")
        self.entry_date.insert(0, date.today().strftime("%d/%m/%Y"))
        self.entry_supplier = ctk.CTkEntry(form_panel, placeholder_text="Fournisseur")
        self.entry_reference = ctk.CTkEntry(form_panel, placeholder_text="Référence produit")
        self.entry_product_combo.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.entry_date.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        self.entry_supplier.grid(row=0, column=2, padx=10, pady=12, sticky="ew")
        self.entry_reference.grid(row=0, column=3, padx=10, pady=12, sticky="ew")

        self.entry_designation = ctk.CTkEntry(form_panel, placeholder_text="Désignation produit")
        self.entry_quantity = ctk.CTkEntry(form_panel, placeholder_text="Quantité")
        self.entry_store = ctk.CTkComboBox(form_panel, values=self._get_store_names())
        self.entry_store.set(self._get_store_names()[0] if self._get_store_names() else "")
        self.entry_designation.grid(row=1, column=0, padx=10, pady=(0, 16), sticky="ew")
        self.entry_quantity.grid(row=1, column=1, padx=10, pady=(0, 16), sticky="ew")
        self.entry_store.grid(row=1, column=2, padx=10, pady=(0, 16), sticky="ew")

        self.entry_add_button = ctk.CTkButton(form_panel, text="Ajouter entrée", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.add_stock_entry)
        self.entry_add_button.grid(row=1, column=3, padx=8, pady=(0, 14), sticky="ew")
        self.entry_save_button = ctk.CTkButton(form_panel, text="Enregistrer entrée", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.save_stock_entry, fg_color="#10B981", hover_color="#059669")
        self.entry_save_button.grid(row=1, column=4, padx=8, pady=(0, 14), sticky="ew")
        self.entry_save_button.configure(state="disabled")
        self.entry_cancel_button = ctk.CTkButton(form_panel, text="Annuler", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.cancel_edit_entry, fg_color="#9CA3AF", hover_color="#6B7280")
        self.entry_cancel_button.grid(row=1, column=5, padx=8, pady=(0, 14), sticky="ew")

        self.entries_search_entry = ctk.CTkEntry(form_panel, placeholder_text="Rechercher entrées...")
        self.entries_search_entry.grid(row=2, column=0, columnspan=4, padx=10, pady=(0, 12), sticky="ew")
        self.entries_search_entry.bind("<KeyRelease>", lambda e: self.search_entries())
        self.entry_delete_button = ctk.CTkButton(form_panel, text="Supprimer sélection", corner_radius=8, height=28, font=ctk.CTkFont(size=11), fg_color="#E53E3E", hover_color="#D53F3F", command=self.delete_selected_entry)
        self.entry_delete_button.grid(row=2, column=4, columnspan=2, padx=8, pady=(0, 12), sticky="ew")

        self.entries_table_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        self.entries_table_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))
        self.entries_table_frame.grid_rowconfigure(0, weight=1)
        self.entries_table_frame.grid_columnconfigure(0, weight=1)

        self.entries_table = ttk.Treeview(self.entries_table_frame, columns=("id", "date", "supplier", "ref", "designation", "store", "qty"), show="headings", selectmode="browse")
        for col, text, width in [("id", "ID", 40), ("date", "Date", 100), ("supplier", "Fournisseur", 160), ("ref", "Référence", 120), ("designation", "Désignation", 220), ("store", "Magasin", 140), ("qty", "Quantité", 90)]:
            self.entries_table.heading(col, text=text)
            self.entries_table.column(col, width=width, anchor="center" if col in ("id", "qty") else "w")
        self.entries_table.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        self.entries_table.bind("<Double-1>", self._load_entry_for_edit)

        scrollbar = ttk.Scrollbar(self.entries_table_frame, orient="vertical", command=self.entries_table.yview)
        self.entries_table.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=8)
        self.entries_status = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.entries_status.grid(row=4, column=0, padx=16, sticky="w")

        self.refresh_entries()

    def _build_outputs(self):
        title = ctk.CTkLabel(self.content, text="Sorties", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Gérer les sorties de stock", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        form_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        form_panel.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        form_panel.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self.output_product_values = self._get_product_references()
        self.output_product_var = tk.StringVar(value=self.output_product_values[0] if self.output_product_values else "")
        self.output_product_combo = ttk.Combobox(form_panel, textvariable=self.output_product_var, values=self.output_product_values, state="normal")
        if self.output_product_values:
            self.output_product_combo.set(self.output_product_values[0])
        self.output_product_combo.bind("<<ComboboxSelected>>", lambda e: self.on_output_product_selected(self.output_product_var.get()))
        self.output_product_combo.bind("<KeyRelease>", self._on_output_product_typed)
        self.output_date = ctk.CTkEntry(form_panel, placeholder_text="Date (jj/mm/aaaa)")
        self.output_date.insert(0, date.today().strftime("%d/%m/%Y"))
        self.output_reference = ctk.CTkEntry(form_panel, placeholder_text="Référence produit")
        self.output_designation = ctk.CTkEntry(form_panel, placeholder_text="Désignation produit")
        self.output_product_combo.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.output_date.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        self.output_reference.grid(row=0, column=2, padx=10, pady=12, sticky="ew")
        self.output_designation.grid(row=0, column=3, padx=10, pady=12, sticky="ew")

        self.output_invoice = ctk.CTkEntry(form_panel, placeholder_text="Facture")
        self.output_destination = ctk.CTkEntry(form_panel, placeholder_text="Destination")
        self.output_quantity = ctk.CTkEntry(form_panel, placeholder_text="Quantité")
        self.output_store = ctk.CTkComboBox(form_panel, values=self._get_store_names())
        self.output_store.set(self._get_store_names()[0] if self._get_store_names() else "")
        self.output_invoice.grid(row=1, column=0, padx=10, pady=(0, 16), sticky="ew")
        self.output_destination.grid(row=1, column=1, padx=10, pady=(0, 16), sticky="ew")
        self.output_quantity.grid(row=1, column=2, padx=10, pady=(0, 16), sticky="ew")
        self.output_store.grid(row=1, column=3, padx=10, pady=(0, 16), sticky="ew")

        self.output_add_button = ctk.CTkButton(form_panel, text="Ajouter sortie", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.add_stock_output)
        self.output_add_button.grid(row=2, column=0, columnspan=2, padx=8, pady=(0, 14), sticky="ew")
        self.output_save_button = ctk.CTkButton(form_panel, text="Enregistrer sortie", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.save_stock_output, fg_color="#10B981", hover_color="#059669")
        self.output_save_button.grid(row=2, column=2, columnspan=2, padx=8, pady=(0, 14), sticky="ew")
        self.output_save_button.configure(state="disabled")
        self.output_cancel_button = ctk.CTkButton(form_panel, text="Annuler", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.cancel_edit_output, fg_color="#9CA3AF", hover_color="#6B7280")
        self.output_cancel_button.grid(row=2, column=4, columnspan=2, padx=8, pady=(0, 14), sticky="ew")

        self.outputs_search_entry = ctk.CTkEntry(form_panel, placeholder_text="Rechercher sorties...")
        self.outputs_search_entry.grid(row=3, column=0, columnspan=4, padx=10, pady=(0, 12), sticky="ew")
        self.outputs_search_entry.bind("<KeyRelease>", lambda e: self.search_outputs())
        self.output_delete_button = ctk.CTkButton(form_panel, text="Supprimer sélection", corner_radius=8, height=28, font=ctk.CTkFont(size=11), fg_color="#E53E3E", hover_color="#D53F3F", command=self.delete_selected_output)
        self.output_delete_button.grid(row=3, column=4, columnspan=2, padx=8, pady=(0, 12), sticky="ew")

        self.outputs_table_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        self.outputs_table_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))
        self.outputs_table_frame.grid_rowconfigure(0, weight=1)
        self.outputs_table_frame.grid_columnconfigure(0, weight=1)

        self.outputs_table = ttk.Treeview(self.outputs_table_frame, columns=("id", "date", "invoice", "ref", "designation", "store", "destination", "qty"), show="headings", selectmode="browse")
        for col, text, width in [("id", "ID", 40), ("date", "Date", 100), ("invoice", "Facture", 120), ("ref", "Référence", 120), ("designation", "Désignation", 180), ("store", "Magasin", 120), ("destination", "Destination", 140), ("qty", "Quantité", 70)]:
            self.outputs_table.heading(col, text=text)
            self.outputs_table.column(col, width=width, anchor="center" if col in ("id", "qty") else "w")
        self.outputs_table.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        self.outputs_table.bind("<Double-1>", self._load_output_for_edit)

        scrollbar = ttk.Scrollbar(self.outputs_table_frame, orient="vertical", command=self.outputs_table.yview)
        self.outputs_table.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=8)
        self.outputs_status = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.outputs_status.grid(row=4, column=0, padx=16, sticky="w")

        self.refresh_outputs()

    def _build_reports(self):
        title = ctk.CTkLabel(self.content, text="Rapports", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Rapports de stock", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        total_products = self.session.query(Product).count()
        total_entries = self.session.query(StockEntry).count()
        total_outputs = self.session.query(StockOutput).count()
        total_entry_qty = self.session.query(func.coalesce(func.sum(StockEntry.quantity), 0)).scalar() or 0
        total_output_qty = self.session.query(func.coalesce(func.sum(StockOutput.quantity), 0)).scalar() or 0

        cards = [
            ("Produits", total_products),
            ("Entrées", total_entries),
            ("Sorties", total_outputs),
            ("Quantité net", total_entry_qty - total_output_qty),
        ]
        report_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        report_frame.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        report_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        for idx, (label, value) in enumerate(cards):
            self._build_stat_card(report_frame, label, value, 0, idx)

        report_button = ctk.CTkButton(report_frame, text="Rapport de tous les produits", corner_radius=10, height=28, font=ctk.CTkFont(size=11), command=self.export_products_pdf)
        report_button.grid(row=1, column=0, columnspan=4, padx=16, pady=(0, 12), sticky="ew")

        history_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        history_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))
        history_frame.grid_rowconfigure(0, weight=1)
        history_frame.grid_columnconfigure(0, weight=1)

        history_label = ctk.CTkLabel(history_frame, text="Historique des 5 dernières transactions", font=ctk.CTkFont(size=14, weight="bold"))
        history_label.grid(row=0, column=0, sticky="w", padx=16, pady=(12, 8))

        history_table = ttk.Treeview(history_frame, columns=("type", "date", "reference", "designation", "qty"), show="headings", selectmode="none")
        headings = [("type", "Type", 100), ("date", "Date", 100), ("reference", "Référence", 120), ("designation", "Désignation", 240), ("qty", "Quantité", 100)]
        for col, text, width in headings:
            history_table.heading(col, text=text)
            history_table.column(col, width=width, anchor="center" if col in ("qty",) else "w")
        history_table.grid(row=1, column=0, sticky="nsew", padx=8, pady=8)

        items = []
        entries = self.session.query(StockEntry).order_by(StockEntry.id.desc()).limit(3).all()
        outputs = self.session.query(StockOutput).order_by(StockOutput.id.desc()).limit(2).all()
        for entry in entries:
            items.append(("Entrée", entry.date, entry.reference, entry.designation, entry.quantity))
        for output in outputs:
            items.append(("Sortie", output.date, output.reference, output.designation, output.quantity))
        for item in items:
            history_table.insert("", "end", values=item)

    def _build_transactions(self):
        title = ctk.CTkLabel(self.content, text="Transactions", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Journal des transactions", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        search_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        search_panel.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 12))
        search_panel.grid_columnconfigure((0, 1), weight=1)

        self.transaction_product_values = ["Tous les produits"] + self._get_product_labels()
        self.transaction_product_var = tk.StringVar(value="Tous les produits")
        self.transaction_product_combo = ttk.Combobox(
            search_panel,
            textvariable=self.transaction_product_var,
            values=self.transaction_product_values,
            state="normal"
        )
        self.transaction_product_combo.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.transaction_product_combo.bind("<<ComboboxSelected>>", lambda event: self.search_transactions())
        self.transaction_product_combo.bind("<KeyRelease>", self._on_transaction_product_typed)

        self.transaction_report_button = ctk.CTkButton(search_panel, text="Rapport transaction", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.export_filtered_transactions_pdf)
        self.transaction_report_button.grid(row=0, column=1, padx=8, pady=10, sticky="ew")

        table_frame = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        table_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.transactions_table = ttk.Treeview(table_frame, columns=("reference", "date", "designation", "store", "in_qty", "out_qty", "balance"), show="headings", selectmode="browse")
        self.transactions_table.configure(displaycolumns=("date", "designation", "store", "in_qty", "out_qty", "balance"))
        columns = [
            ("date", "Date", 120),
            ("designation", "Désignation", 260),
            ("store", "Magasin", 160),
            ("in_qty", "Quantité entrée", 120),
            ("out_qty", "Quantité sortie", 120),
            ("balance", "Stock", 100),
        ]
        for col, text, width in columns:
            self.transactions_table.heading(col, text=text)
            self.transactions_table.column(col, width=width, anchor="center" if col in ("in_qty", "out_qty", "balance") else "w")
        self.transactions_table.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        self.transactions_table.bind("<<TreeviewSelect>>", self.on_transaction_row_selected)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.transactions_table.yview)
        self.transactions_table.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=8)

        self.transaction_status = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.transaction_status.grid(row=4, column=0, padx=16, sticky="w")

        self.refresh_transactions()

    def _build_settings(self):
        title = ctk.CTkLabel(self.content, text="Paramètres", font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(8, 0), padx=16)
        subtitle = ctk.CTkLabel(self.content, text="Configuration de l'application", font=ctk.CTkFont(size=12))
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        settings = self._get_company_settings()
        form_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        form_panel.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        form_panel.grid_columnconfigure((0, 1), weight=1)

        self.settings_name = ctk.CTkEntry(form_panel, placeholder_text="Nom de l'entreprise")
        self.settings_address = ctk.CTkEntry(form_panel, placeholder_text="Adresse")
        self.settings_city = ctk.CTkEntry(form_panel, placeholder_text="Ville")
        self.settings_phone = ctk.CTkEntry(form_panel, placeholder_text="Téléphone")
        self.settings_email = ctk.CTkEntry(form_panel, placeholder_text="Email")
        self.settings_website = ctk.CTkEntry(form_panel, placeholder_text="Site web")
        self.settings_tax_id = ctk.CTkEntry(form_panel, placeholder_text="Numéro fisc")
        self.settings_rccm = ctk.CTkEntry(form_panel, placeholder_text="RCCM")

        self.settings_name.grid(row=0, column=0, padx=10, pady=12, sticky="ew")
        self.settings_address.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        self.settings_city.grid(row=1, column=0, padx=10, pady=12, sticky="ew")
        self.settings_phone.grid(row=1, column=1, padx=10, pady=12, sticky="ew")
        self.settings_email.grid(row=2, column=0, padx=10, pady=12, sticky="ew")
        self.settings_website.grid(row=2, column=1, padx=10, pady=12, sticky="ew")
        self.settings_tax_id.grid(row=3, column=0, padx=10, pady=12, sticky="ew")
        self.settings_rccm.grid(row=3, column=1, padx=10, pady=12, sticky="ew")

        self.settings_save_button = ctk.CTkButton(self.content, text="Enregistrer paramètres", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.save_company_settings)
        self.settings_save_button.grid(row=3, column=0, padx=16, pady=(0, 12), sticky="w")

        self.settings_status = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.settings_status.grid(row=4, column=0, padx=16, sticky="w")

        users_panel = ctk.CTkFrame(self.content, corner_radius=12, fg_color="#0D1117")
        users_panel.grid(row=5, column=0, sticky="ew", padx=16, pady=(0, 16))
        users_panel.grid_columnconfigure((0, 1, 2), weight=1)

        user_title = ctk.CTkLabel(users_panel, text="Créer un utilisateur magasinier", font=ctk.CTkFont(size=14, weight="bold"))
        user_title.grid(row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(12, 8))

        self.new_user_username = ctk.CTkEntry(users_panel, placeholder_text="Nom d'utilisateur")
        self.new_user_password = ctk.CTkEntry(users_panel, placeholder_text="Mot de passe", show="*")
        self.new_user_username.grid(row=1, column=0, padx=10, pady=(0, 12), sticky="ew")
        self.new_user_password.grid(row=1, column=1, padx=10, pady=(0, 12), sticky="ew")

        self.create_user_button = ctk.CTkButton(users_panel, text="Créer magasinier", corner_radius=8, height=28, font=ctk.CTkFont(size=11), command=self.create_magasinier_user)
        self.create_user_button.grid(row=1, column=2, padx=8, pady=(0, 12), sticky="ew")

        self.user_management_status = ctk.CTkLabel(self.content, text="", anchor="w", text_color="#8b95a1")
        self.user_management_status.grid(row=6, column=0, padx=16, sticky="w")

        self._load_settings(settings)

    def _get_store_names(self):
        stores = [store.name for store in self.session.query(Store).order_by(Store.name).all()]
        return stores or ["Aucun magasin"]

    def _get_product_labels(self):
        products = self.session.query(Product).order_by(Product.reference).all()
        labels = [f"{product.reference} — {product.designation}" for product in products]
        return labels or ["Aucun produit"]

    def _get_product_references(self):
        products = self.session.query(Product).order_by(Product.reference).all()
        refs = [product.reference for product in products]
        return refs or ["Aucun produit"]

    def _normalize_text(self, text):
        if text is None:
            return ""
        s = str(text).strip().lower()
        s = unicodedata.normalize("NFD", s)
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        return s
    def _split_product_label(self, label):
        if not label or "—" not in label:
            return "", ""
        reference, designation = label.split("—", 1)
        return reference.strip(), designation.strip()

    def _load_product_combo(self):
        values = self._get_product_labels()
        # Entry combobox
        if hasattr(self, "entry_product_combo"):
            refs = self._get_product_references()
            self.entry_product_values = refs
            self.entry_product_combo.configure(values=refs)
            if refs:
                self.entry_product_combo.set(refs[0])
        # Output combobox
        if hasattr(self, "output_product_combo"):
            refs = self._get_product_references()
            self.output_product_values = refs
            self.output_product_combo.configure(values=refs)
            if refs:
                self.output_product_combo.set(refs[0])
        # Transaction combobox
        if hasattr(self, "transaction_product_combo"):
            values_with_all = ["Tous les produits"] + values
            self.transaction_product_values = values_with_all
            self.transaction_product_combo.configure(values=values_with_all)
            if values_with_all:
                self.transaction_product_combo.set(values_with_all[0])

    def on_entry_product_selected(self, reference):
        if not reference:
            return
        self.entry_reference.delete(0, tk.END)
        self.entry_reference.insert(0, reference)
        product = self.session.query(Product).filter_by(reference=reference).first()
        self.entry_designation.delete(0, tk.END)
        if product:
            self.entry_designation.insert(0, product.designation)
            if product.store_id:
                store = self.session.get(Store, product.store_id)
                if store:
                    self.entry_store.set(store.name)
        else:
            self.entry_designation.insert(0, "")

    def on_output_product_selected(self, reference):
        if not reference:
            return
        self.output_reference.delete(0, tk.END)
        self.output_reference.insert(0, reference)
        product = self.session.query(Product).filter_by(reference=reference).first()
        self.output_designation.delete(0, tk.END)
        if product:
            self.output_designation.insert(0, product.designation)
            if product.store_id:
                store = self.session.get(Store, product.store_id)
                if store:
                    self.output_store.configure(values=[store.name])
                    self.output_store.set(store.name)
                    self.output_store.configure(state="disabled")
                else:
                    self.output_store.configure(values=self._get_store_names(), state="normal")
                    if self._get_store_names():
                        self.output_store.set(self._get_store_names()[0])
            else:
                self.output_store.configure(values=self._get_store_names(), state="normal")
                if self._get_store_names():
                    self.output_store.set(self._get_store_names()[0])
        else:
            self.output_designation.insert(0, "")
            self.output_store.configure(values=self._get_store_names(), state="normal")
            if self._get_store_names():
                self.output_store.set(self._get_store_names()[0])

    def _on_entry_product_typed(self, event=None):
        typed = self.entry_product_var.get().strip()
        if not typed:
            values = getattr(self, 'entry_product_values', self._get_product_references())
        else:
            q = self._normalize_text(typed)
            values = [v for v in getattr(self, 'entry_product_values', []) if q in self._normalize_text(v)]
            if not values:
                values = getattr(self, 'entry_product_values', self._get_product_references())
        self.entry_product_combo.configure(values=values)
        if typed in values:
            self.on_entry_product_selected(typed)

    def _on_output_product_typed(self, event=None):
        typed = self.output_product_var.get().strip()
        if not typed:
            values = getattr(self, 'output_product_values', self._get_product_references())
        else:
            q = self._normalize_text(typed)
            values = [v for v in getattr(self, 'output_product_values', []) if q in self._normalize_text(v)]
            if not values:
                values = getattr(self, 'output_product_values', self._get_product_references())
        self.output_product_combo.configure(values=values)
        if typed in values:
            self.on_output_product_selected(typed)

    def _parse_date(self, date_str):
        try:
            return datetime.strptime(date_str, "%d/%m/%Y")
        except Exception:
            return datetime.max

    def _get_transaction_rows(self, selected_ref=None, query=None):
        query = self._normalize_text(query or "")
        rows = []
        if selected_ref:
            entries = self.session.query(StockEntry, Store.name).join(Store, StockEntry.store_id == Store.id).filter(StockEntry.reference == selected_ref).all()
            outputs = self.session.query(StockOutput, Store.name).join(Store, StockOutput.store_id == Store.id).filter(StockOutput.reference == selected_ref).all()
            for e, store_name in entries:
                rows.append({
                    "type": "Entrée",
                    "date": e.date,
                    "reference": e.reference,
                    "designation": e.designation,
                    "store": store_name or "—",
                    "in_qty": e.quantity,
                    "out_qty": 0,
                })
            for o, store_name in outputs:
                rows.append({
                    "type": "Sortie",
                    "date": o.date,
                    "reference": o.reference,
                    "designation": o.designation,
                    "store": store_name or "—",
                    "in_qty": 0,
                    "out_qty": o.quantity,
                })
            rows.sort(key=lambda r: (self._parse_date(r["date"]), r["type"]))
            product = self.session.query(Product).filter_by(reference=selected_ref).first()
            balance = product.initial_stock if product else 0
            detailed = []
            for row in rows:
                if query:
                    text = f"{row['date']} {row['reference']} {row['designation']} {row['store']} {row['in_qty']} {row['out_qty']}"
                    if query not in self._normalize_text(text):
                        continue
                balance += row["in_qty"] - row["out_qty"]
                row["balance"] = balance
                detailed.append(row)
            return detailed

        entries = self.session.query(StockEntry, Store.name).join(Store, StockEntry.store_id == Store.id).all()
        outputs = self.session.query(StockOutput, Store.name).join(Store, StockOutput.store_id == Store.id).all()
        for e, store_name in entries:
            rows.append({
                "type": "Entrée",
                "date": e.date,
                "reference": e.reference,
                "designation": e.designation,
                "store": store_name or "—",
                "in_qty": e.quantity,
                "out_qty": 0,
            })
        for o, store_name in outputs:
            rows.append({
                "type": "Sortie",
                "date": o.date,
                "reference": o.reference,
                "designation": o.designation,
                "store": store_name or "—",
                "in_qty": 0,
                "out_qty": o.quantity,
            })
        rows.sort(key=lambda r: (self._parse_date(r["date"]), r["reference"], r["type"]))

        initial_stocks = {p.reference: p.initial_stock or 0 for p in self.session.query(Product).all()}
        balances = {}
        detailed = []
        for row in rows:
            if query:
                text = f"{row['date']} {row['reference']} {row['designation']} {row['store']} {row['in_qty']} {row['out_qty']}"
                if query not in self._normalize_text(text):
                    continue
            ref = row["reference"]
            if ref not in balances:
                balances[ref] = initial_stocks.get(ref, 0)
            balances[ref] += row["in_qty"] - row["out_qty"]
            row["balance"] = balances[ref]
            detailed.append(row)
        return detailed

    def search_transactions(self, event=None):
        selected_input = self.transaction_product_var.get().strip()
        normalized_input = self._normalize_text(selected_input)
        selected_ref = None
        rows = []

        if not selected_input or selected_input == "Tous les produits":
            rows = self._get_transaction_rows(None, None)
        elif selected_input in self.transaction_product_values:
            selected_ref, _ = self._split_product_label(selected_input)
            rows = self._get_transaction_rows(selected_ref, None)
        else:
            matching_refs = [self._split_product_label(label)[0] for label in self.transaction_product_values[1:] if normalized_input in self._normalize_text(label)]
            if matching_refs:
                all_rows = self._get_transaction_rows(None, None)
                rows = [row for row in all_rows if row["reference"] in matching_refs]
            else:
                rows = self._get_transaction_rows(None, normalized_input)

        self.transactions_table.delete(*self.transactions_table.get_children())
        for row in rows:
            in_text = f"{row['in_qty']}" if row['in_qty'] else ""
            out_text = f"{row['out_qty']}" if row['out_qty'] else ""
            self.transactions_table.insert("", "end", values=(row['reference'], row['date'], row['designation'], row['store'], in_text, out_text, str(row['balance'])))

        if selected_ref:
            self.transaction_status.configure(text=f"{len(rows)} transaction(s) pour {selected_ref}")
        elif selected_input and selected_input != "Tous les produits":
            self.transaction_status.configure(text=f"{len(rows)} transaction(s) correspondant à '{selected_input}'")
        else:
            self.transaction_status.configure(text=f"{len(rows)} produit(s) affiché(s)")

    def _on_transaction_product_typed(self, event=None):
        typed = self.transaction_product_var.get().strip()
        if not typed:
            values = self.transaction_product_values
        else:
            query = self._normalize_text(typed)
            values = [value for value in self.transaction_product_values if query in self._normalize_text(value)]
            if not values:
                values = self.transaction_product_values
        self.transaction_product_combo.configure(values=values)
        self.search_transactions()

    def export_filtered_transactions_pdf(self):
        if not REPORTLAB_AVAILABLE:
            message = "reportlab est requis pour exporter en PDF."
            print(f"[CustomTkinter] {message}")
            messagebox.showerror("Export impossible", message)
            return
        selected_product = self.transaction_product_var.get().strip()
        if not selected_product or selected_product == "Tous les produits":
            messagebox.showinfo("Produit requis", "Sélectionnez d'abord un produit pour générer le rapport transactionnel.")
            return
        if selected_product not in self.transaction_product_values:
            messagebox.showinfo("Produit invalide", "Sélectionnez un produit existant dans la liste pour générer le rapport.")
            return
        selected_ref, _ = self._split_product_label(selected_product)
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        transactions = self._get_transaction_rows(selected_ref, None)
        if not transactions:
            messagebox.showinfo("Aucun résultat", "Aucune transaction à exporter.")
            return
        self._build_transactions_pdf(path, transactions, selected_product)
        messagebox.showinfo("Export réussi", f"Fichier PDF enregistré :\n{path}")

    def _build_transactions_pdf(self, path, transactions, query=None):
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        title_text = "Rapport des transactions"
        if query:
            title_text += f" - {query}"
        title_style = ParagraphStyle(name="Title", fontName="Helvetica-Bold", fontSize=18, textColor=colors.HexColor("#111827"))

        data = [["Date", "Produit", "Magasin", "Entrées", "Sorties", "Stock"]]
        for transaction in transactions:
            if isinstance(transaction, dict):
                product_text = f"{transaction.get('reference', '')} — {transaction.get('designation', '')}".strip(' —')
                data.append([
                    transaction.get('date', ''),
                    product_text,
                    transaction.get('store', ''),
                    str(transaction.get('in_qty', '')) if transaction.get('in_qty', 0) else "",
                    str(transaction.get('out_qty', '')) if transaction.get('out_qty', 0) else "",
                    str(transaction.get('balance', '')),
                ])
            else:
                # fallback for list/tuple rows
                data.append(list(transaction))

        table = Table(data, colWidths=[35*mm, 80*mm, 50*mm, 35*mm, 35*mm, 30*mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ]))
        story = [Paragraph(title_text, title_style), Spacer(1, 8), table]
        doc.build(story)

    def _get_company_settings(self):
        settings = self.session.get(CompanySettings, 1)
        if not settings:
            settings = CompanySettings(id=1)
            self.session.add(settings)
            self.session.commit()
        return settings

    def _load_settings(self, settings):
        self.settings_name.delete(0, tk.END)
        self.settings_address.delete(0, tk.END)
        self.settings_city.delete(0, tk.END)
        self.settings_phone.delete(0, tk.END)
        self.settings_email.delete(0, tk.END)
        self.settings_website.delete(0, tk.END)
        self.settings_tax_id.delete(0, tk.END)
        self.settings_rccm.delete(0, tk.END)

        self.settings_name.insert(0, settings.name or "")
        self.settings_address.insert(0, settings.address or "")
        self.settings_city.insert(0, settings.city or "")
        self.settings_phone.insert(0, settings.phone or "")
        self.settings_email.insert(0, settings.email or "")
        self.settings_website.insert(0, settings.website or "")
        self.settings_tax_id.insert(0, settings.tax_id or "")
        self.settings_rccm.insert(0, settings.rccm or "")

    def save_company_settings(self):
        settings = self._get_company_settings()
        settings.name = self.settings_name.get().strip()
        settings.address = self.settings_address.get().strip()
        settings.city = self.settings_city.get().strip()
        settings.phone = self.settings_phone.get().strip()
        settings.email = self.settings_email.get().strip()
        settings.website = self.settings_website.get().strip()
        settings.tax_id = self.settings_tax_id.get().strip()
        settings.rccm = self.settings_rccm.get().strip()
        self.session.add(settings)
        self.session.commit()
        self.settings_status.configure(text="Paramètres enregistrés.")

    def create_magasinier_user(self):
        if self.current_role != "admin":
            self.user_management_status.configure(text="Seul l'administrateur peut créer un magasinier.")
            return

        username = self.new_user_username.get().strip()
        password = self.new_user_password.get().strip()
        if not username or not password:
            self.user_management_status.configure(text="Nom d'utilisateur et mot de passe requis.")
            return

        if self.session.query(User).filter(func.lower(User.username) == username.lower()).first():
            self.user_management_status.configure(text="Ce nom d'utilisateur existe déjà.")
            return

        salt, password_hash = self._hash_password(password)
        user = User(username=username, password_hash=password_hash, password_salt=salt, role="magasinier")
        self.session.add(user)
        self.session.commit()
        self.new_user_username.delete(0, tk.END)
        self.new_user_password.delete(0, tk.END)
        self.user_management_status.configure(text=f"Magasinier {username} créé avec succès.")

    def add_stock_entry(self):
        date_value = self.entry_date.get().strip()
        supplier = self.entry_supplier.get().strip()
        reference = self.entry_reference.get().strip()
        designation = self.entry_designation.get().strip()
        store_name = self.entry_store.get().strip()
        quantity = self.entry_quantity.get().strip()

        if not date_value or not reference or not designation or not store_name or not quantity:
            messagebox.showwarning("Valeurs manquantes", "Tous les champs sont obligatoires pour une entrée.")
            return
        try:
            quantity_value = int(quantity)
        except ValueError:
            messagebox.showwarning("Quantité invalide", "Entrez un nombre entier pour la quantité.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        entry = StockEntry(date=date_value, supplier=supplier, reference=reference, designation=designation, store_id=store.id, quantity=quantity_value)
        self.session.add(entry)
        self.session.commit()
        self.entries_status.configure(text=f"Entrée {reference} ajoutée.")
        self.refresh_entries()

    def _load_entry_for_edit(self, event=None):
        selected = self.entries_table.selection()
        if not selected:
            messagebox.showinfo("Sélection vide", "Sélectionnez une entrée pour modifier.")
            return
        entry_id = int(self.entries_table.item(selected[0], "values")[0])
        entry = self.session.get(StockEntry, entry_id)
        if not entry:
            messagebox.showerror("Erreur", "Entrée non trouvée.")
            return

        store = self.session.get(Store, entry.store_id) if entry.store_id else None
        self.entry_product_combo.set(entry.reference)
        self.entry_date.delete(0, tk.END)
        self.entry_date.insert(0, entry.date)
        self.entry_supplier.delete(0, tk.END)
        self.entry_supplier.insert(0, entry.supplier or "")
        self.entry_reference.delete(0, tk.END)
        self.entry_reference.insert(0, entry.reference)
        self.entry_designation.delete(0, tk.END)
        self.entry_designation.insert(0, entry.designation)
        self.entry_quantity.delete(0, tk.END)
        self.entry_quantity.insert(0, str(entry.quantity))
        if store:
            self.entry_store.set(store.name)
        else:
            self.entry_store.set(self._get_store_names()[0] if self._get_store_names() else "")

        self.editing_entry_id = entry_id
        self.entry_add_button.configure(state="disabled")
        self.entry_save_button.configure(state="normal")
        self.entry_cancel_button.configure(state="normal")
        self.entries_status.configure(text=f"✎ Mode édition: entrée {entry.reference}")

    def save_stock_entry(self):
        if self.editing_entry_id is None:
            messagebox.showinfo("Aucune modification", "Sélectionnez une entrée puis cliquez sur Modifier pour enregistrer les changements.")
            return

        date_value = self.entry_date.get().strip()
        supplier = self.entry_supplier.get().strip()
        reference = self.entry_reference.get().strip()
        designation = self.entry_designation.get().strip()
        store_name = self.entry_store.get().strip()
        quantity = self.entry_quantity.get().strip()

        if not date_value or not reference or not designation or not store_name or not quantity:
            messagebox.showwarning("Valeurs manquantes", "Tous les champs sont obligatoires pour une entrée.")
            return
        try:
            quantity_value = int(quantity)
        except ValueError:
            messagebox.showwarning("Quantité invalide", "Entrez un nombre entier pour la quantité.")
            return

        entry = self.session.get(StockEntry, self.editing_entry_id)
        if not entry:
            messagebox.showerror("Erreur", "Entrée non trouvée pour la modification.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        entry.date = date_value
        entry.supplier = supplier
        entry.reference = reference
        entry.designation = designation
        entry.store_id = store.id
        entry.quantity = quantity_value
        self.session.commit()

        self.cancel_edit_entry()
        self.entries_status.configure(text=f"✓ Entrée {reference} modifiée avec succès.")
        self.refresh_entries()

    def cancel_edit_entry(self):
        if self.editing_entry_id is None:
            return
        self.editing_entry_id = None
        refs = self._get_product_references()
        self.entry_product_combo.set(refs[0] if refs else "")
        self.entry_date.delete(0, tk.END)
        self.entry_date.insert(0, date.today().strftime("%d/%m/%Y"))
        self.entry_supplier.delete(0, tk.END)
        self.entry_reference.delete(0, tk.END)
        self.entry_designation.delete(0, tk.END)
        self.entry_quantity.delete(0, tk.END)
        self.entry_store.set(self._get_store_names()[0] if self._get_store_names() else "")
        self.entry_add_button.configure(state="normal")
        self.entry_save_button.configure(state="disabled")
        self.entry_cancel_button.configure(state="normal")
        self.entries_status.configure(text="")
        self.entries_table.selection_remove(self.entries_table.selection())

    def search_entries(self):
        query = self._normalize_text(self.entries_search_entry.get() if hasattr(self, "entries_search_entry") else "")
        self.entries_table.delete(*self.entries_table.get_children())
        entries = self.session.query(StockEntry, Store.name).outerjoin(Store, StockEntry.store_id == Store.id).order_by(StockEntry.id.desc()).all()
        count = 0
        for entry, store_name in entries:
            if query:
                text = f"{entry.date} {entry.supplier} {entry.reference} {entry.designation} {store_name or ''}"
                if query not in self._normalize_text(text):
                    continue
            self.entries_table.insert("", "end", values=(entry.id, entry.date, entry.supplier, entry.reference, entry.designation, store_name or "—", entry.quantity))
            count += 1
        if hasattr(self, "entries_status"):
            self.entries_status.configure(text=f"{count} entrée(s) affichée(s)")

    def delete_selected_entry(self):
        selected = self.entries_table.selection()
        if not selected:
            messagebox.showinfo("Sélection vide", "Sélectionnez une entrée à supprimer.")
            return
        entry_id = int(self.entries_table.item(selected[0], "values")[0])
        entry = self.session.get(StockEntry, entry_id)
        if not entry:
            messagebox.showerror("Erreur", "Entrée non trouvée.")
            return
        if messagebox.askyesno("Confirmation", f"Supprimer l'entrée {entry.reference} du {entry.date} ?"):
            self.session.delete(entry)
            self.session.commit()
            self.cancel_edit_entry()
            self.search_entries()
            self.entries_status.configure(text="Entrée supprimée.")

    def refresh_entries(self):
        if hasattr(self, "entries_search_entry"):
            self.search_entries()
        else:
            self.entries_table.delete(*self.entries_table.get_children())
            entries = self.session.query(StockEntry, Store.name).outerjoin(Store, StockEntry.store_id == Store.id).order_by(StockEntry.id.desc()).all()
            for entry, store_name in entries:
                self.entries_table.insert("", "end", values=(entry.id, entry.date, entry.supplier, entry.reference, entry.designation, store_name or "—", entry.quantity))

    def add_stock_output(self):
        date_value = self.output_date.get().strip()
        reference = self.output_reference.get().strip()
        designation = self.output_designation.get().strip()
        invoice = self.output_invoice.get().strip()
        destination = self.output_destination.get().strip()
        store_name = self.output_store.get().strip()
        quantity = self.output_quantity.get().strip()

        if not date_value or not reference or not designation or not store_name or not quantity:
            messagebox.showwarning("Valeurs manquantes", "Tous les champs sont obligatoires pour une sortie.")
            return
        try:
            quantity_value = int(quantity)
        except ValueError:
            messagebox.showwarning("Quantité invalide", "Entrez un nombre entier pour la quantité.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        output = StockOutput(date=date_value, reference=reference, designation=designation, invoice_number=invoice, destination=destination, store_id=store.id, quantity=quantity_value)
        self.session.add(output)
        self.session.commit()
        self.outputs_status.configure(text=f"Sortie {reference} ajoutée.")
        self.refresh_outputs()

    def _load_output_for_edit(self, event=None):
        selected = self.outputs_table.selection()
        if not selected:
            messagebox.showinfo("Sélection vide", "Sélectionnez une sortie pour modifier.")
            return
        output_id = int(self.outputs_table.item(selected[0], "values")[0])
        output = self.session.get(StockOutput, output_id)
        if not output:
            messagebox.showerror("Erreur", "Sortie non trouvée.")
            return

        store = self.session.get(Store, output.store_id) if output.store_id else None
        self.output_product_combo.set(output.reference)
        self.output_date.delete(0, tk.END)
        self.output_date.insert(0, output.date)
        self.output_reference.delete(0, tk.END)
        self.output_reference.insert(0, output.reference)
        self.output_designation.delete(0, tk.END)
        self.output_designation.insert(0, output.designation)
        self.output_invoice.delete(0, tk.END)
        self.output_invoice.insert(0, output.invoice_number or "")
        self.output_destination.delete(0, tk.END)
        self.output_destination.insert(0, output.destination or "")
        self.output_quantity.delete(0, tk.END)
        self.output_quantity.insert(0, str(output.quantity))
        if store:
            self.output_store.configure(values=[store.name], state="disabled")
            self.output_store.set(store.name)
        else:
            self.output_store.configure(values=self._get_store_names(), state="normal")
            self.output_store.set(self._get_store_names()[0] if self._get_store_names() else "")

        self.editing_output_id = output_id
        self.output_add_button.configure(state="disabled")
        self.output_save_button.configure(state="normal")
        self.output_cancel_button.configure(state="normal")
        self.outputs_status.configure(text=f"✎ Mode édition: sortie {output.reference}")

    def save_stock_output(self):
        if self.editing_output_id is None:
            messagebox.showinfo("Aucune modification", "Sélectionnez une sortie puis cliquez sur Modifier pour enregistrer les changements.")
            return

        date_value = self.output_date.get().strip()
        reference = self.output_reference.get().strip()
        designation = self.output_designation.get().strip()
        invoice = self.output_invoice.get().strip()
        destination = self.output_destination.get().strip()
        store_name = self.output_store.get().strip()
        quantity = self.output_quantity.get().strip()

        if not date_value or not reference or not designation or not store_name or not quantity:
            messagebox.showwarning("Valeurs manquantes", "Tous les champs sont obligatoires pour une sortie.")
            return
        try:
            quantity_value = int(quantity)
        except ValueError:
            messagebox.showwarning("Quantité invalide", "Entrez un nombre entier pour la quantité.")
            return

        output = self.session.get(StockOutput, self.editing_output_id)
        if not output:
            messagebox.showerror("Erreur", "Sortie non trouvée pour la modification.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        output.date = date_value
        output.reference = reference
        output.designation = designation
        output.invoice_number = invoice
        output.destination = destination
        output.store_id = store.id
        output.quantity = quantity_value
        self.session.commit()

        self.cancel_edit_output()
        self.outputs_status.configure(text=f"✓ Sortie {reference} modifiée avec succès.")
        self.refresh_outputs()

    def save_stock(self):
        if self._current_page == 2:
            self.save_stock_entry()
        elif self._current_page == 3:
            self.save_stock_output()
        else:
            messagebox.showinfo("Aucune action", "Aucune opération d'enregistrement disponible pour cette page.")

    def cancel_edit_output(self):
        if self.editing_output_id is None:
            return
        self.editing_output_id = None
        refs = self._get_product_references()
        self.output_product_combo.set(refs[0] if refs else "")
        self.output_date.delete(0, tk.END)
        self.output_date.insert(0, date.today().strftime("%d/%m/%Y"))
        self.output_reference.delete(0, tk.END)
        self.output_designation.delete(0, tk.END)
        self.output_invoice.delete(0, tk.END)
        self.output_destination.delete(0, tk.END)
        self.output_quantity.delete(0, tk.END)
        self.output_store.configure(values=self._get_store_names(), state="normal")
        if self._get_store_names():
            self.output_store.set(self._get_store_names()[0])
        self.output_add_button.configure(state="normal")
        self.output_save_button.configure(state="disabled")
        self.output_cancel_button.configure(state="normal")
        self.outputs_status.configure(text="")
        self.outputs_table.selection_remove(self.outputs_table.selection())

    def search_outputs(self):
        query = self._normalize_text(self.outputs_search_entry.get() if hasattr(self, "outputs_search_entry") else "")
        self.outputs_table.delete(*self.outputs_table.get_children())
        outputs = self.session.query(StockOutput, Store.name).outerjoin(Store, StockOutput.store_id == Store.id).order_by(StockOutput.id.desc()).all()
        count = 0
        for output, store_name in outputs:
            if query:
                text = f"{output.date} {output.invoice_number} {output.reference} {output.designation} {store_name or ''} {output.destination}"
                if query not in self._normalize_text(text):
                    continue
            self.outputs_table.insert("", "end", values=(output.id, output.date, output.invoice_number, output.reference, output.designation, store_name or "—", output.destination, output.quantity))
            count += 1
        if hasattr(self, "outputs_status"):
            self.outputs_status.configure(text=f"{count} sortie(s) affichée(s)")

    def delete_selected_output(self):
        selected = self.outputs_table.selection()
        if not selected:
            messagebox.showinfo("Sélection vide", "Sélectionnez une sortie à supprimer.")
            return
        output_id = int(self.outputs_table.item(selected[0], "values")[0])
        output = self.session.get(StockOutput, output_id)
        if not output:
            messagebox.showerror("Erreur", "Sortie non trouvée.")
            return
        if messagebox.askyesno("Confirmation", f"Supprimer la sortie {output.reference} du {output.date} ?"):
            self.session.delete(output)
            self.session.commit()
            self.cancel_edit_output()
            self.search_outputs()
            self.outputs_status.configure(text="Sortie supprimée.")

    def refresh_outputs(self):
        # Reset the store combo to normal state with all stores
        self.output_store.configure(values=self._get_store_names(), state="normal")
        if self._get_store_names():
            self.output_store.set(self._get_store_names()[0])
        if hasattr(self, "outputs_search_entry"):
            self.search_outputs()
        else:
            self.outputs_table.delete(*self.outputs_table.get_children())
            outputs = self.session.query(StockOutput, Store.name).outerjoin(Store, StockOutput.store_id == Store.id).order_by(StockOutput.id.desc()).all()
            for output, store_name in outputs:
                self.outputs_table.insert("", "end", values=(output.id, output.date, output.invoice_number, output.reference, output.designation, store_name or "—", output.destination, output.quantity))

    def refresh_transactions(self):
        self.transactions_table.delete(*self.transactions_table.get_children())
        rows = self._get_transaction_rows(None, None)
        for row in rows:
            in_text = f"{row['in_qty']}" if row['in_qty'] else ""
            out_text = f"{row['out_qty']}" if row['out_qty'] else ""
            self.transactions_table.insert("", "end", values=(row['reference'], row['date'], row['designation'], row['store'], in_text, out_text, str(row['balance'])))
        if hasattr(self, "transaction_status"):
            self.transaction_status.configure(text=f"{len(rows)} produit(s) affiché(s)")

    def on_transaction_row_selected(self, event):
        selected = self.transactions_table.selection()
        if not selected:
            return
        row_values = self.transactions_table.item(selected[0], "values")
        if not row_values:
            return
        reference = row_values[0]
        product = self.session.query(Product).filter_by(reference=reference).first()
        if not product:
            return
        label = f"{product.reference} — {product.designation}"
        values = ["Tous les produits"] + self._get_product_labels()
        if label not in values:
            return
        self.transaction_product_combo.set(label)
        self.search_transactions()

    def _get_product_current_stocks(self):
        """Return dict mapping product reference -> actual current stock."""
        entry_sums = dict(
            self.session.query(StockEntry.reference, func.sum(StockEntry.quantity))
            .group_by(StockEntry.reference).all()
        )
        output_sums = dict(
            self.session.query(StockOutput.reference, func.sum(StockOutput.quantity))
            .group_by(StockOutput.reference).all()
        )
        return {
            p.reference: (p.initial_stock or 0) + entry_sums.get(p.reference, 0) - output_sums.get(p.reference, 0)
            for p in self.session.query(Product).all()
        }

    def _build_stat_card(self, parent, title, value, row, col):
        card = ctk.CTkFrame(parent, corner_radius=12, fg_color="#0D1117")
        card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card_title = ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=12))
        card_title.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
        card_value = ctk.CTkLabel(card, text=str(value), font=ctk.CTkFont(size=24, weight="bold"))
        card_value.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))
        return card

    def refresh_products(self):
        if self._current_page != 1:
            return
        self._load_stores()
        self.table.delete(*self.table.get_children())
        current_stocks = self._get_product_current_stocks()
        products = self.session.query(Product, Store.name).outerjoin(Store, Product.store_id == Store.id).order_by(Product.reference).all()
        for idx, (product, store_name) in enumerate(products):
            stock = current_stocks.get(product.reference, product.initial_stock or 0)
            tag = "lowstock" if stock <= 0 else ("evenrow" if idx % 2 else "oddrow")
            self.table.insert("", "end", iid=str(product.id), values=(product.id, product.reference, product.designation, product.unit, stock, store_name or "—"), tags=(tag,))
        self.status_label.configure(text=f"{len(products)} produit(s) affiché(s)")

    def _load_stores(self):
        values = self._get_store_names()
        self.store_combo.configure(values=values)
        if values:
            self.store_combo.set(values[0])

    def add_product(self):
        ref = self.ref_entry.get().strip()
        des = self.des_entry.get().strip()
        unit = self.unit_entry.get().strip() or "unité"
        store_name = self.store_combo.get().strip()
        try:
            initial_stock = int(self.stock_entry.get().strip() or 0)
        except ValueError:
            initial_stock = 0

        if not ref or not des or not store_name:
            messagebox.showwarning("Valeurs manquantes", "Référence, désignation et magasin sont obligatoires.")
            return

        if self.session.query(Product).filter_by(reference=ref).first():
            messagebox.showwarning("Référence existante", f"La référence {ref} existe déjà.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        product = Product(reference=ref, designation=des, unit=unit, initial_stock=initial_stock, store_id=store.id)
        self.session.add(product)
        self.session.commit()
        self.ref_entry.delete(0, tk.END)
        self.des_entry.delete(0, tk.END)
        self.unit_entry.delete(0, tk.END)
        self.unit_entry.insert(0, "unité")
        self.stock_entry.delete(0, tk.END)
        self.refresh_products()
        self._load_product_combo()
        self.status_label.configure(text=f"✓ Produit {des} ajouté avec succès.")

    def save_product(self):
        if self.editing_product_id is None:
            messagebox.showinfo("Aucune modification", "Sélectionnez un produit puis cliquez sur Modifier pour enregistrer les changements.")
            return

        ref = self.ref_entry.get().strip()
        des = self.des_entry.get().strip()
        unit = self.unit_entry.get().strip() or "unité"
        store_name = self.store_combo.get().strip()
        try:
            initial_stock = int(self.stock_entry.get().strip() or 0)
        except ValueError:
            initial_stock = 0

        if not ref or not des or not store_name:
            messagebox.showwarning("Valeurs manquantes", "Référence, désignation et magasin sont obligatoires.")
            return

        product = self.session.get(Product, self.editing_product_id)
        if not product:
            messagebox.showerror("Erreur", "Produit non trouvé pour la modification.")
            return

        if ref != product.reference and self.session.query(Product).filter_by(reference=ref).first():
            messagebox.showwarning("Référence existante", f"La référence {ref} existe déjà.")
            return

        store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
        if not store:
            store = Store(name=store_name)
            self.session.add(store)
            self.session.commit()
            self._load_stores()

        product.reference = ref
        product.designation = des
        product.unit = unit
        product.initial_stock = initial_stock
        product.store_id = store.id
        self.session.commit()

        self.ref_entry.delete(0, tk.END)
        self.des_entry.delete(0, tk.END)
        self.unit_entry.delete(0, tk.END)
        self.unit_entry.insert(0, "unité")
        self.stock_entry.delete(0, tk.END)
        self.store_combo.set(self._get_store_names()[0] if self._get_store_names() else "")

        self.editing_product_id = None
        self.add_button.configure(text="+ Ajouter", state="normal")
        self.save_button.configure(state="disabled")
        self.modify_button.configure(text="Modifier")

        self.refresh_products()
        self._load_product_combo()
        self.status_label.configure(text=f"✓ Produit {des} modifié avec succès.")

    def delete_selected(self):
        selection = self.table.selection()
        if not selection:
            messagebox.showinfo("Sélection vide", "Sélectionnez au moins un produit à supprimer.")
            return
        products = []
        for item_id in selection:
            product = self.session.get(Product, int(item_id))
            if product:
                products.append(product)
        if not products:
            return
        if messagebox.askyesno("Confirmation", f"Supprimer {len(products)} produit(s) sélectionné(s) ?"):
            for product in products:
                self.session.delete(product)
            self.session.commit()
            self.refresh_products()
            self._load_product_combo()
            self.status_label.configure(text=f"{len(products)} produit(s) supprimé(s).")

    def select_all_products(self):
        items = self.table.get_children()
        if not items:
            return
        self.table.selection_set(items)
        self.status_label.configure(text=f"{len(items)} produit(s) sélectionné(s).")

    def cancel_edit_product(self):
        """Annuler le mode édition et réinitialiser les champs."""
        if self.editing_product_id is None:
            return
        
        self.editing_product_id = None
        self.ref_entry.delete(0, tk.END)
        self.des_entry.delete(0, tk.END)
        self.unit_entry.delete(0, tk.END)
        self.unit_entry.insert(0, "unité")
        self.stock_entry.delete(0, tk.END)
        self.store_combo.set(self._get_store_names()[0] if self._get_store_names() else "")
        self.add_button.configure(text="+ Ajouter", state="normal")
        self.save_button.configure(state="disabled")
        self.modify_button.configure(text="Modifier")
        self.status_label.configure(text="")
        self.table.selection_remove(self.table.selection())

    def modify_selected_product(self):
        """Charger le produit sélectionné pour édition dans le formulaire."""
        self._edit_product_from_table()
    def _edit_product_from_table(self, event=None):
        """Charger les données du produit sélectionné pour édition dans le formulaire."""
        item = self.table.selection()
        if not item:
            messagebox.showinfo("Sélection vide", "Sélectionnez un produit à modifier.")
            return
        product_id = int(self.table.item(item[0], "values")[0])
        product = self.session.get(Product, product_id)
        if not product:
            messagebox.showerror("Erreur", "Produit non trouvé.")
            return

        # Remplir les champs du formulaire avec les données du produit sélectionné
        self.ref_entry.delete(0, tk.END)
        self.ref_entry.insert(0, product.reference)

        self.des_entry.delete(0, tk.END)
        self.des_entry.insert(0, product.designation)

        self.unit_entry.delete(0, tk.END)
        self.unit_entry.insert(0, product.unit or "unité")

        self.stock_entry.delete(0, tk.END)
        self.stock_entry.insert(0, str(product.initial_stock))

        # Sélectionner le magasin associé
        store = self.session.get(Store, product.store_id) if product.store_id else None
        if store:
            # Remplir le combo du magasin avec la valeur correspondante
            self.store_combo.set(store.name)
        else:
            self.store_combo.set("")

        # Stocker l'ID du produit actuellement édité
        self.editing_product_id = product_id

        # Passage en mode édition
        self.add_button.configure(text="+ Ajouter", state="disabled")
        self.save_button.configure(state="normal")
        self.modify_button.configure(text="Modifier")
        self.status_label.configure(text=f"✎ Mode édition: {product.reference} — cliquez sur Enregistrer pour sauvegarder")

        # Focus sur le premier champ
        self.ref_entry.focus()

    def search_products(self):
        raw = self.search_entry.get()
        query = self._normalize_text(raw)
        current_stocks = self._get_product_current_stocks()

        self.table.delete(*self.table.get_children())
        products = self.session.query(Product, Store.name).outerjoin(Store, Product.store_id == Store.id).order_by(Product.reference).all()
        visible_count = 0
        for idx, (product, store_name) in enumerate(products):
            text = f"{product.reference} {product.designation} {store_name or ''}"
            if query and query not in self._normalize_text(text):
                continue
            stock = current_stocks.get(product.reference, product.initial_stock or 0)
            tag = "lowstock" if stock <= 0 else ("evenrow" if visible_count % 2 else "oddrow")
            self.table.insert("", "end", iid=str(product.id), values=(product.id, product.reference, product.designation, product.unit, stock, store_name or "—"), tags=(tag,))
            visible_count += 1

        if query:
            self.status_label.configure(text=f"{visible_count} résultat(s)")
        else:
            self.status_label.configure(text=f"{visible_count} produit(s) affiché(s)")

    def import_products(self):
        if not OPENPYXL_AVAILABLE:
            message = f"openpyxl est requis pour importer Excel.\nInterpréteur: {sys.executable}"
            print(f"[CustomTkinter] {message}")
            messagebox.showerror("Import impossible", message)
            return
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if not path:
            return

        def normalize_header(value):
            if value is None:
                return ""
            text = str(value).strip().lower()
            text = unicodedata.normalize("NFD", text)
            text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
            return text.replace(" ", "").replace("-", "").replace("_", "")

        ref_keys = {"reference", "ref", "referenc", "refeference"}
        des_keys = {"designation", "designatio", "libelle", "label"}
        unit_keys = {"unite", "unit"}
        stock_keys = {"stockinitial", "stockinitiale", "initialstock", "stock"}
        store_keys = {"magasin", "store"}

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Le fichier Excel est vide.")

            headers = [normalize_header(cell) for cell in rows[0]]
            header_map = {h: idx for idx, h in enumerate(headers) if h}

            ref_header = next((h for h in headers if h in ref_keys), None)
            des_header = next((h for h in headers if h in des_keys), None)
            if ref_header is None or des_header is None:
                raise ValueError("Les colonnes référence et désignation sont requises.")

            imported = skipped = 0
            for row in rows[1:]:
                if not row:
                    continue
                ref = row[header_map.get(ref_header)] if ref_header is not None else None
                des = row[header_map.get(des_header)] if des_header is not None else None
                if ref is None or des is None:
                    continue
                ref = str(ref).strip()
                des = str(des).strip()
                if not ref or not des:
                    continue
                if self.session.query(Product).filter_by(reference=ref).first():
                    skipped += 1
                    continue

                unit = "unité"
                unit_header = next((h for h in headers if h in unit_keys), None)
                if unit_header is not None and header_map.get(unit_header, 0) < len(row):
                    unit_value = row[header_map[unit_header]]
                    if unit_value is not None:
                        unit = str(unit_value).strip() or unit

                initial_stock = 0
                stock_header = next((h for h in headers if h in stock_keys), None)
                if stock_header is not None and header_map.get(stock_header, 0) < len(row):
                    stock_val = row[header_map[stock_header]]
                    if stock_val is not None:
                        try:
                            initial_stock = int(stock_val)
                        except Exception:
                            initial_stock = 0

                store_name = None
                store_header = next((h for h in headers if h in store_keys), None)
                if store_header is not None and header_map.get(store_header, 0) < len(row):
                    store_value = row[header_map[store_header]]
                    if store_value is not None:
                        store_name = str(store_value).strip()

                store = None
                if store_name:
                    store = self.session.query(Store).filter(func.lower(Store.name) == store_name.lower()).first()
                    if not store:
                        store = Store(name=store_name)
                        self.session.add(store)
                        self.session.commit()
                if not store:
                    store = self.session.query(Store).order_by(Store.name).first()

                product = Product(reference=ref, designation=des, unit=unit, initial_stock=initial_stock, store_id=store.id if store else None)
                self.session.add(product)
                imported += 1
            self.session.commit()
            self.refresh_products()
            messagebox.showinfo("Import réussi", f"{imported} produit(s) importé(s).\n{skipped} produit(s) ignoré(s).")
        except Exception as exc:
            print(f"[CustomTkinter] Import exception: {exc}")
            messagebox.showerror("Erreur import", str(exc))

    def export_products_pdf(self):
        if not REPORTLAB_AVAILABLE:
            message = "reportlab est requis pour exporter en PDF."
            print(f"[CustomTkinter] {message}")
            messagebox.showerror("Export impossible", message)
            return
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not path:
            return
        products = [(self.table.item(item, "values")) for item in self.table.get_children()]
        try:
            self._build_pdf(path, products)
            messagebox.showinfo("Export réussi", f"Fichier PDF enregistré :\n{path}")
        except Exception as exc:
            print(f"[CustomTkinter] Export exception: {exc}")
            messagebox.showerror("Erreur export", str(exc))

    def _build_pdf(self, path, products):
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        title_style = ParagraphStyle(name="Title", fontName="Helvetica-Bold", fontSize=18, textColor=colors.HexColor("#111827"))
        normal_style = ParagraphStyle(name="Normal", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#1f2937"))
        data = [["ID", "Référence", "Désignation", "Unité", "Stock", "Magasin"]]
        for product in products:
            data.append(list(product))
        table = Table(data, colWidths=[25*mm, 40*mm, 110*mm, 25*mm, 25*mm, 50*mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ]))
        story = [Paragraph("Catalogue des produits", title_style), Spacer(1, 8), table]
        doc.build(story)


def run_customtk_app():
    app = ProductManagerApp()
    app.mainloop()
